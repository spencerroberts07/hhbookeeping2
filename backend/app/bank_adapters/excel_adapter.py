"""Excel (.xlsx/.xls) adapter — header-mapped deterministic parse via openpyxl.

Detects a header row, maps columns by fuzzy name (date / description /
amount, or debit + credit), and emits signed NormalizedBankTxn rows. Same
column-mapping idea the CSV importer uses, so one mental model covers both.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl

from .base import AdapterResult, NormalizedBankTxn, StatementMeta

_DATE_KEYS = ("date", "posted", "transaction date", "trans date")
_DESC_KEYS = ("description", "details", "narrative", "memo", "transaction")
_AMOUNT_KEYS = ("amount", "value")
_DEBIT_KEYS = ("debit", "withdrawal", "withdrawals", "paid out", "out")
_CREDIT_KEYS = ("credit", "deposit", "deposits", "paid in", "in")
_BAL_KEYS = ("balance", "running balance")


def _match(header: str, keys) -> bool:
    h = (header or "").strip().lower()
    return any(k in h for k in keys)


def _to_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%b %d, %Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _to_dec(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v).replace(",", "").replace("$", "").strip())
    except (InvalidOperation, ValueError):
        return None


class ExcelBankAdapter:
    source_system = "statement_excel"

    def supports(self, *, filename: str, content_type: str, sniff: bytes) -> bool:
        return (filename or "").lower().endswith((".xlsx", ".xlsm", ".xls")) or sniff[:2] == b"PK"

    def parse(self, *, file_bytes: bytes, filename: str, entity_code: str,
              hints: dict[str, Any] | None = None) -> AdapterResult:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # find the header row (first row mapping a date + (amount or debit/credit))
        hdr_idx, colmap = None, {}
        for i, row in enumerate(rows[:15]):
            cells = [str(c).strip() if c is not None else "" for c in row]
            m: dict[str, int] = {}
            for j, c in enumerate(cells):
                if "date" not in m and _match(c, _DATE_KEYS): m["date"] = j
                if "desc" not in m and _match(c, _DESC_KEYS): m["desc"] = j
                if "amount" not in m and _match(c, _AMOUNT_KEYS): m["amount"] = j
                if "debit" not in m and _match(c, _DEBIT_KEYS): m["debit"] = j
                if "credit" not in m and _match(c, _CREDIT_KEYS): m["credit"] = j
                if "balance" not in m and _match(c, _BAL_KEYS): m["balance"] = j
            if "date" in m and ("amount" in m or ("debit" in m and "credit" in m)):
                hdr_idx, colmap = i, m
                break
        if hdr_idx is None:
            return AdapterResult(meta=StatementMeta(source_system=self.source_system),
                                 transactions=[], parser_confidence=0.0,
                                 warnings=["Excel: no recognizable header row (date + amount/debit/credit)"])

        txns: list[NormalizedBankTxn] = []
        closing: Decimal | None = None
        for row in rows[hdr_idx + 1:]:
            def cell(key):
                j = colmap.get(key)
                return row[j] if j is not None and j < len(row) else None
            d = _to_date(cell("date"))
            if d is None:
                continue
            if "amount" in colmap:
                amt = _to_dec(cell("amount"))
            else:
                deb, cred = _to_dec(cell("debit")), _to_dec(cell("credit"))
                amt = (cred or Decimal("0")) - (deb or Decimal("0"))
            if amt is None:
                continue
            bal = _to_dec(cell("balance"))
            if bal is not None:
                closing = bal
            txns.append(NormalizedBankTxn(
                transaction_date=d, amount=amt,
                description=str(cell("desc") or "").strip(),
                running_balance=bal,
                raw={"row": [str(c) if c is not None else None for c in row]},
            ))

        meta = StatementMeta(
            source_account_code=hints.get("source_account_code") if hints else None,
            closing_balance=closing,
            statement_date=txns[-1].transaction_date if txns else None,
            source_system=self.source_system,
        )
        warnings = [] if txns else ["Excel: header found but no data rows parsed"]
        return AdapterResult(meta=meta, transactions=txns, parser_confidence=1.0, warnings=warnings)
