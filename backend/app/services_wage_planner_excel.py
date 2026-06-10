"""
Wage Cost Planner — Excel workbook generator.

Produces a three-sheet workbook:
  1. Inputs / Assumptions    — annual settings + salaried roster
  2. Results                 — 26-period tracker table (mirrors the dealer's
                               manual Excel tracker layout)
  3. Min-Wage Calculator     — impact of a proposed minimum-wage increase

Uses openpyxl.  Falls back gracefully with a ValueError if openpyxl is not
installed (same pattern as services_depreciation.py).

NOT called directly by routes — routes call generate_wage_planner_excel() to
get bytes, then upload via storage_service (services_storage.py).
"""
from __future__ import annotations

import logging
from datetime import date
from decimal import Decimal
from typing import Any
from uuid import UUID

from .services_wage_planner import compute_plan, get_settings, min_wage_impact

_log = logging.getLogger(__name__)


def _fmt_dec(v) -> str:
    if v is None:
        return ""
    return str(Decimal(str(v)).quantize(Decimal("0.01")))


def _fmt_pct(v) -> str:
    if v is None:
        return ""
    return f"{float(v) * 100:.2f}%"


def generate_wage_planner_excel(
    session,
    *,
    entity_id: UUID,
    fiscal_year: int,
) -> bytes:
    """Return a wage-planner workbook as bytes.
    Raises ValueError if openpyxl is not installed or settings are missing."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError("openpyxl is required for wage-planner Excel export") from exc

    plan = compute_plan(session, entity_id=entity_id, fiscal_year=fiscal_year)
    if not plan["settings"]:
        raise ValueError(
            f"No wage planner settings for entity {entity_id} FY{fiscal_year}. "
            "Configure settings before downloading."
        )

    settings_row = plan["settings"]
    periods = plan["periods"]
    summary = plan["summary"]

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Sheet 1 — Inputs / Assumptions
    # ------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Inputs & Assumptions"

    _bold = Font(bold=True)
    _title_font = Font(bold=True, size=12)
    _header_fill = PatternFill("solid", fgColor="1F4E79")
    _header_font = Font(bold=True, color="FFFFFF", size=10)
    _currency = '#,##0.00'
    _pct_fmt = '0.00%'

    ws1["A1"] = f"Wage Cost Planner — FY{fiscal_year}"
    ws1["A1"].font = _title_font
    ws1.merge_cells("A1:C1")

    rows_a = [
        ("", ""),
        ("Annual Assumptions", "Value"),
        ("Target Wage % of Sales", settings_row["target_wage_pct"]),
        ("Forecast Sales Change vs Prior Year", settings_row["forecast_sales_change"]),
        ("Avg Hourly Wage (excl. salaried)", settings_row["avg_hourly_wage"]),
        ("Benefits %", settings_row["benefits_pct"]),
        ("Distribution Basis", settings_row["distribution_basis"].replace("_", " ").title()),
        ("Notes", settings_row.get("notes") or ""),
        ("", ""),
        ("Forecast Annual Sales", summary.get("forecast_annual_sales")),
        ("Target Annual Wage Cost $", summary.get("target_annual_wage_dollars")),
        ("", ""),
        ("Salaried Staff", "Annual Salary", "Bonus", "Hours/Period", "Wage/Period"),
    ]
    for r in rows_a:
        ws1.append(list(r))

    # Salaried staff detail
    salaried = settings_row.get("salaried_staff") or []
    bp = Decimal(str(settings_row["benefits_pct"]))
    B = Decimal("1") + bp
    for emp in salaried:
        sal = Decimal(str(emp["annual_salary"])) + Decimal(str(emp["bonus"]))
        wage_pp = float(sal * B / Decimal("26"))
        ws1.append([
            emp["employee_name"],
            emp["annual_salary"],
            emp["bonus"],
            emp.get("assumed_hours_per_period", 80),
            round(wage_pp, 2),
        ])

    # Style header row (row 3 = index 3)
    for cell in ws1["3"]:
        cell.font = _bold

    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 14

    # ------------------------------------------------------------------
    # Sheet 2 — Results (26-period tracker)
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Results")

    ws2["A1"] = f"Wage Cost Planner — FY{fiscal_year} — Period Results"
    ws2["A1"].font = _title_font
    ws2.merge_cells("A1:M1")

    headers2 = [
        "Period #",
        "Period Start",
        "Period End",
        "LY Sales",
        "Forecast Sales",
        "Target Wage $",
        "Target Hours",
        "Actual Sales",
        "Actual Gross Wages",
        "Actual Stat Pay",
        "Actual Hours",
        "Hours Over/(Under)",
        "Adj. Go-Fwd Target Hrs",
        "Actual $/Hr",
        "LY $/Hr",
        "Locked",
    ]
    ws2.append([""])  # blank row 2 (title merged on row 1)
    ws2.append(headers2)

    header_row_idx = 3
    for cell in ws2[header_row_idx]:
        cell.fill = _header_fill
        cell.font = _header_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
    ws2.row_dimensions[header_row_idx].height = 28

    for p in periods:
        row = [
            p["period_number"],
            p["period_start"].isoformat() if isinstance(p["period_start"], date) else str(p.get("period_start") or ""),
            p["period_end"].isoformat() if isinstance(p["period_end"], date) else str(p.get("period_end") or ""),
            float(p["py_sales"]) if p["py_sales"] is not None else None,
            float(p["forecast_sales"]) if p["forecast_sales"] is not None else None,
            float(p["target_wage_dollars"]) if p["target_wage_dollars"] is not None else None,
            float(p["target_hours"]) if p["target_hours"] is not None else None,
            float(p["actual_sales"]) if p["actual_sales"] is not None else None,
            float(p["actual_gross_wages"]) if p["actual_gross_wages"] is not None else None,
            float(p["actual_stat_pay"]) if p["actual_stat_pay"] is not None else None,
            float(p["actual_hours"]) if p["actual_hours"] is not None else None,
            float(p["hours_over_under"]) if p["hours_over_under"] is not None else None,
            float(p["adjusted_target_hours"]) if p["adjusted_target_hours"] is not None else None,
            float(p["actual_sales_per_hour"]) if p["actual_sales_per_hour"] is not None else None,
            float(p["py_sales_per_hour"]) if p["py_sales_per_hour"] is not None else None,
            "Yes" if p["locked"] else "",
        ]
        ws2.append(row)

    # Apply currency format to dollar columns (D, E, F, H, I, J, N, O)
    dollar_cols = [4, 5, 6, 8, 9, 10, 14, 15]
    hour_cols = [7, 11, 12, 13]
    start_row = header_row_idx + 1
    end_row = start_row + len(periods) - 1
    for col in dollar_cols:
        for r in range(start_row, end_row + 1):
            ws2.cell(row=r, column=col).number_format = _currency
    for col in hour_cols:
        for r in range(start_row, end_row + 1):
            ws2.cell(row=r, column=col).number_format = '#,##0.00'

    # Summary row
    ws2.append([])
    sum_row = [
        "TOTAL", "", "",
        float(summary["forecast_annual_sales"]) if summary.get("forecast_annual_sales") else None,
        float(summary["forecast_annual_sales"]) if summary.get("forecast_annual_sales") else None,
        float(summary["target_annual_wage_dollars"]) if summary.get("target_annual_wage_dollars") else None,
    ]
    ws2.append(sum_row)
    total_row_idx = end_row + 3
    for cell in ws2[total_row_idx]:
        cell.font = Font(bold=True)

    col_widths2 = [8, 12, 12, 14, 14, 14, 13, 14, 16, 14, 13, 16, 18, 12, 12, 8]
    for i, w in enumerate(col_widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------
    # Sheet 3 — Min-Wage Calculator (uses live current roster)
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Min-Wage Calculator")

    ws3["A1"] = "Minimum Wage Impact Calculator"
    ws3["A1"].font = _title_font
    ws3.merge_cells("A1:F1")

    ws3.append([])
    ws3.append([
        "Employee",
        "Current Rate",
        "New Min. Rate",
        "Delta Rate",
        "Current Bi-weekly Est.",
        "Projected Bi-weekly Est.",
        "Delta Bi-weekly Est.",
    ])
    for cell in ws3[3]:
        cell.fill = _header_fill
        cell.font = _header_font

    # Default scenario: Ontario 2026 minimum wage $17.20
    try:
        impact = min_wage_impact(session, entity_id=entity_id, new_min_wage=17.20)
        for emp in impact["employees"]:
            ws3.append([
                emp["full_name"],
                float(emp["current_rate"]),
                float(emp["new_rate"]),
                float(emp["delta_rate"]),
                float(emp["current_biweekly_est"]),
                float(emp["projected_biweekly_est"]),
                float(emp["delta_biweekly_est"]),
            ])
        ws3.append([])
        ws3.append([
            f"TOTAL ({impact['affected_employees']} employees affected at $17.20)",
            "", "", "",
            float(impact["total_current_biweekly_est"]),
            float(impact["total_projected_biweekly_est"]),
            float(impact["total_delta_biweekly_est"]),
        ])
        ws3.append([
            "Annual delta estimate",
            "", "", "", "", "",
            float(impact["total_delta_annual_est"]),
        ])
    except Exception as exc:
        _log.warning("wage_planner excel: min-wage sheet error: %r", exc)
        ws3.append([f"Error computing min-wage impact: {exc}"])

    for col in [2, 3, 4, 5, 6, 7]:
        for r in range(4, ws3.max_row + 1):
            ws3.cell(row=r, column=col).number_format = _currency

    ws3.column_dimensions["A"].width = 30
    for i in range(2, 8):
        ws3.column_dimensions[get_column_letter(i)].width = 18

    # ------------------------------------------------------------------
    # Serialize
    # ------------------------------------------------------------------
    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
