"""
Fixed asset / depreciation — service layer.

Canadian CCA declining-balance method:
    annual_dep = opening_NBV * cca_rate
    half-year rule (year of acquisition only): annual_dep *= 0.5

Bridlewood asset classes (seeded by seed_fixed_assets):
    Equipment / Furniture & Fixtures  Class 8   15%   GL 1510 / 1610 / 6900
    Computer Equipment                Class 8   15%   GL 1540 / 1640 / 6900
    Vehicles                          Class 10  30%   GL 1520 / 1620 / 6900

Depreciation expense rolls up to a single P&L account (6900); only
accumulated-depn (1610/1620/1640) splits per class on the balance sheet.

New in Module B (migration 052):
    • fixed_asset_classes  — per-entity class config table
    • fixed_asset_disposals — disposal event records
    • compute_monthly_depreciation_by_class() — primary interface for Module A
    • generate_excel_schedule()               — downloadable Excel schedule
    • post_disposal_journal()                 — disposal JE (requires gain/loss
                                                accounts; stub until CoA confirmed)
"""
from __future__ import annotations

import json
from datetime import date
from decimal import Decimal
from typing import Any
from uuid import UUID

from sqlalchemy import text

from .services import (
    _has_table,
    _parse_uuid,
    get_entity_by_code,
    get_or_create_accounting_period,
)


SOURCE_MODULE_DEPRECIATION = "depreciation"
BATCH_LABEL_DEPRECIATION = "monthly_depreciation"

CCA_CLASS_8_EQUIPMENT = "class_8_equipment"
CCA_CLASS_8_COMPUTER = "class_8_computer"
CCA_CLASS_10_VEHICLE = "class_10_vehicle"

# Bridlewood collapses depreciation expense into a single P&L account.
# Accumulated depreciation stays split per class on the balance sheet.
DEPN_EXPENSE_GL_ACCOUNT = "6900"

# Canonical Bridlewood seed list (driven from the actual fixed-asset
# schedule the user supplied). opening_nbv_date = fiscal year start.
_BRIDLEWOOD_SEED = [
    {
        "asset_code": "EQUIP-001",
        "description": "Equipment / Furniture & Fixtures",
        "cca_class": CCA_CLASS_8_EQUIPMENT,
        "cca_rate": Decimal("0.15"),
        "asset_gl_account": "1510",
        "accum_depn_gl_account": "1610",
        "depn_expense_gl_account": DEPN_EXPENSE_GL_ACCOUNT,
        "cost": Decimal("386378.00"),
        "opening_nbv": Decimal("357400.00"),
        "opening_nbv_date": date(2024, 9, 30),
    },
    {
        "asset_code": "COMP-001",
        "description": "Computer Equipment",
        "cca_class": CCA_CLASS_8_COMPUTER,
        "cca_rate": Decimal("0.15"),
        "asset_gl_account": "1540",
        "accum_depn_gl_account": "1640",
        "depn_expense_gl_account": DEPN_EXPENSE_GL_ACCOUNT,
        "cost": Decimal("13098.00"),
        "opening_nbv": Decimal("12116.00"),
        "opening_nbv_date": date(2024, 9, 30),
    },
    {
        "asset_code": "VEH-001",
        "description": "Vehicles",
        "cca_class": CCA_CLASS_10_VEHICLE,
        "cca_rate": Decimal("0.30"),
        "asset_gl_account": "1520",
        "accum_depn_gl_account": "1620",
        "depn_expense_gl_account": DEPN_EXPENSE_GL_ACCOUNT,
        "cost": Decimal("30000.00"),
        "opening_nbv": Decimal("21000.00"),
        "opening_nbv_date": date(2024, 9, 30),
    },
]


def _money(value: Any) -> Decimal:
    return Decimal(str(value or 0)).quantize(Decimal("0.01"))


# ----------------------------------------------------------------------
# Seed
# ----------------------------------------------------------------------


def seed_fixed_assets(
    session, *, entity_code: str, actor_email: str
) -> dict[str, Any]:
    if not actor_email:
        raise ValueError("actor_email is required")
    entity = get_entity_by_code(session, entity_code)
    if not entity:
        raise ValueError(f"Unknown entity code: {entity_code}")

    inserted = 0
    skipped = 0
    rows = []
    for cfg in _BRIDLEWOOD_SEED:
        existing = session.execute(
            text(
                "SELECT id FROM fixed_assets "
                "WHERE entity_id = :entity_id AND asset_code = :asset_code"
            ),
            {"entity_id": entity["id"], "asset_code": cfg["asset_code"]},
        ).mappings().first()
        if existing:
            skipped += 1
            rows.append({"asset_code": cfg["asset_code"], "status": "exists"})
            continue
        ins = session.execute(
            text(
                """
                INSERT INTO fixed_assets (
                    entity_id, asset_code, description, cca_class, cca_rate,
                    asset_gl_account, accum_depn_gl_account,
                    depn_expense_gl_account,
                    acquisition_date, cost, opening_nbv, opening_nbv_date,
                    is_active, notes
                ) VALUES (
                    :entity_id, :asset_code, :description, :cca_class, :cca_rate,
                    :asset_gl, :accum_gl, :exp_gl,
                    NULL, :cost, :opening_nbv, :opening_nbv_date,
                    TRUE, :notes
                )
                RETURNING id
                """
            ),
            {
                "entity_id": entity["id"],
                "asset_code": cfg["asset_code"],
                "description": cfg["description"],
                "cca_class": cfg["cca_class"],
                "cca_rate": cfg["cca_rate"],
                "asset_gl": cfg["asset_gl_account"],
                "accum_gl": cfg["accum_depn_gl_account"],
                "exp_gl": cfg["depn_expense_gl_account"],
                "cost": cfg["cost"],
                "opening_nbv": cfg["opening_nbv"],
                "opening_nbv_date": cfg["opening_nbv_date"],
                "notes": f"Seeded for {entity_code} by {actor_email}",
            },
        ).mappings().first()
        inserted += 1
        rows.append(
            {
                "asset_code": cfg["asset_code"],
                "id": str(ins["id"]),
                "status": "inserted",
            }
        )

    return {
        "entity_code": entity_code,
        "inserted": inserted,
        "skipped": skipped,
        "assets": rows,
    }


# ----------------------------------------------------------------------
# Depreciation math
# ----------------------------------------------------------------------


def calculate_annual_depreciation(
    opening_nbv: Decimal, cca_rate: Decimal, half_year_rule: bool
) -> Decimal:
    base = Decimal(str(opening_nbv)) * Decimal(str(cca_rate))
    if half_year_rule:
        base = base * Decimal("0.5")
    return _money(base)


def generate_depreciation_schedule(
    session,
    *,
    entity_code: str,
    fiscal_year: int,
    actor_email: str,
    half_year_asset_codes: list[str] | None = None,
) -> dict[str, Any]:
    """
    Compute and store a per-asset annual schedule. half_year_asset_codes
    is the list of asset codes acquired in this fiscal year (the
    half-year rule applies to those only).
    """
    if not actor_email:
        raise ValueError("actor_email is required")
    entity = get_entity_by_code(session, entity_code)
    if not entity:
        raise ValueError(f"Unknown entity code: {entity_code}")

    half_year_codes = set(half_year_asset_codes or [])

    assets = session.execute(
        text(
            """
            SELECT id, asset_code, description, cca_class, cca_rate,
                   opening_nbv, opening_nbv_date,
                   asset_gl_account, accum_depn_gl_account,
                   depn_expense_gl_account
            FROM fixed_assets
            WHERE entity_id = :entity_id AND is_active = TRUE
            ORDER BY asset_code
            """
        ),
        {"entity_id": entity["id"]},
    ).mappings().all()

    if not assets:
        raise ValueError(
            "No active fixed_assets found. Call /api/depreciation/seed-assets first."
        )

    rows_out: list[dict[str, Any]] = []
    total_annual = Decimal("0.00")

    for asset in assets:
        # If a prior fiscal year's schedule exists, the current opening_nbv
        # is the prior year's closing_nbv; otherwise use the asset's
        # opening_nbv from the asset record.
        prior = session.execute(
            text(
                """
                SELECT closing_nbv
                FROM depreciation_schedules
                WHERE entity_id = :entity_id
                  AND fixed_asset_id = :asset_id
                  AND fiscal_year < :fiscal_year
                ORDER BY fiscal_year DESC
                LIMIT 1
                """
            ),
            {
                "entity_id": entity["id"],
                "asset_id": asset["id"],
                "fiscal_year": int(fiscal_year),
            },
        ).mappings().first()
        opening_nbv = _money(
            prior["closing_nbv"] if prior else asset["opening_nbv"]
        )
        rate = Decimal(str(asset["cca_rate"]))
        half_year = asset["asset_code"] in half_year_codes
        annual = calculate_annual_depreciation(opening_nbv, rate, half_year)
        monthly = (annual / Decimal("12")).quantize(Decimal("0.01"))
        closing = _money(opening_nbv - annual)

        session.execute(
            text(
                """
                INSERT INTO depreciation_schedules (
                    entity_id, fixed_asset_id, fiscal_year,
                    opening_nbv, annual_cca_rate, half_year_rule_applies,
                    annual_depreciation, monthly_depreciation, closing_nbv
                ) VALUES (
                    :entity_id, :asset_id, :fiscal_year,
                    :opening_nbv, :rate, :half_year,
                    :annual, :monthly, :closing
                )
                ON CONFLICT (entity_id, fixed_asset_id, fiscal_year)
                DO UPDATE SET
                    opening_nbv = EXCLUDED.opening_nbv,
                    annual_cca_rate = EXCLUDED.annual_cca_rate,
                    half_year_rule_applies = EXCLUDED.half_year_rule_applies,
                    annual_depreciation = EXCLUDED.annual_depreciation,
                    monthly_depreciation = EXCLUDED.monthly_depreciation,
                    closing_nbv = EXCLUDED.closing_nbv
                """
            ),
            {
                "entity_id": entity["id"],
                "asset_id": asset["id"],
                "fiscal_year": int(fiscal_year),
                "opening_nbv": opening_nbv,
                "rate": rate,
                "half_year": half_year,
                "annual": annual,
                "monthly": monthly,
                "closing": closing,
            },
        )

        total_annual += annual
        rows_out.append(
            {
                "asset_code": asset["asset_code"],
                "description": asset["description"],
                "cca_class": asset["cca_class"],
                "cca_rate": str(rate),
                "half_year_rule_applies": half_year,
                "opening_nbv": str(opening_nbv),
                "annual_depreciation": str(annual),
                "monthly_depreciation": str(monthly),
                "closing_nbv": str(closing),
            }
        )

    return {
        "entity_code": entity_code,
        "fiscal_year": int(fiscal_year),
        "asset_count": len(rows_out),
        "total_annual_depreciation": str(total_annual.quantize(Decimal("0.01"))),
        "rows": rows_out,
    }


# ----------------------------------------------------------------------
# Journal builder
# ----------------------------------------------------------------------


def _fiscal_year_for_period(entity: dict[str, Any], period_end: date) -> int:
    """
    Bridlewood's fiscal year ends 9/30 (per entities.fiscal_year_end_*).
    Fiscal year is labelled by the calendar year in which it ends.
    """
    fy_month = entity.get("fiscal_year_end_month") or 9
    fy_day = entity.get("fiscal_year_end_day") or 30
    if (period_end.month, period_end.day) <= (fy_month, fy_day):
        return period_end.year
    return period_end.year + 1


def build_depreciation_journal(
    session,
    *,
    entity_code: str,
    period_end: date,
    actor_email: str,
) -> dict[str, Any]:
    if not actor_email:
        raise ValueError("actor_email is required")
    entity = get_entity_by_code(session, entity_code)
    if not entity:
        raise ValueError(f"Unknown entity code: {entity_code}")

    period_start = period_end.replace(day=1)
    accounting_period_id = get_or_create_accounting_period(
        session, entity["id"], period_end
    )
    fiscal_year = _fiscal_year_for_period(entity, period_end)

    # Pull active assets and their schedule rows for this fiscal year.
    rows = session.execute(
        text(
            """
            SELECT fa.id AS asset_id, fa.asset_code, fa.description,
                   fa.depn_expense_gl_account, fa.accum_depn_gl_account,
                   ds.id AS schedule_id, ds.monthly_depreciation
            FROM fixed_assets fa
            LEFT JOIN depreciation_schedules ds
                ON ds.fixed_asset_id = fa.id
                   AND ds.entity_id = fa.entity_id
                   AND ds.fiscal_year = :fiscal_year
            WHERE fa.entity_id = :entity_id AND fa.is_active = TRUE
            ORDER BY fa.asset_code
            """
        ),
        {"entity_id": entity["id"], "fiscal_year": fiscal_year},
    ).mappings().all()

    if not rows:
        raise ValueError(
            "No active fixed_assets found. Seed assets and generate "
            "the schedule before posting depreciation."
        )
    missing = [r["asset_code"] for r in rows if r["schedule_id"] is None]
    if missing:
        raise ValueError(
            f"No depreciation_schedules row for fiscal_year={fiscal_year} "
            f"on assets: {missing}. Generate the schedule first."
        )

    # Build the journal: one Dr+Cr pair per asset.
    journal_lines: list[dict[str, Any]] = []
    total = Decimal("0.00")
    line_number = 0
    for r in rows:
        monthly = _money(r["monthly_depreciation"])
        if monthly == Decimal("0.00"):
            continue
        total += monthly
        line_number += 1
        journal_lines.append(
            {
                "line_number": line_number,
                "account_code": r["depn_expense_gl_account"],
                "debit_amount": monthly,
                "credit_amount": Decimal("0.00"),
                "memo": f"Monthly depreciation — {r['description']}",
                "asset_id": r["asset_id"],
            }
        )
        line_number += 1
        journal_lines.append(
            {
                "line_number": line_number,
                "account_code": r["accum_depn_gl_account"],
                "debit_amount": Decimal("0.00"),
                "credit_amount": monthly,
                "memo": f"Monthly depreciation — {r['description']}",
                "asset_id": r["asset_id"],
            }
        )

    if total == Decimal("0.00"):
        raise ValueError("All assets have zero monthly depreciation; nothing to post.")

    # Idempotency: refuse if any per-asset journal line already exists for
    # this period. (Combined check; the unique constraint also enforces.)
    already_posted = session.execute(
        text(
            """
            SELECT fixed_asset_id
            FROM depreciation_journal_lines
            WHERE entity_id = :entity_id
              AND period_start = :period_start
            """
        ),
        {"entity_id": entity["id"], "period_start": period_start},
    ).mappings().all()
    if already_posted:
        # Still allow re-posting by overwriting the journal_batches row,
        # but warn the caller via the response.
        existing_count = len(already_posted)
    else:
        existing_count = 0

    summary = {
        "fiscal_year": fiscal_year,
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "asset_count": len(rows),
        "total_monthly_depreciation": str(total.quantize(Decimal("0.01"))),
        "previously_posted_asset_count": existing_count,
    }

    batch = session.execute(
        text(
            """
            INSERT INTO journal_batches (
                entity_id, accounting_period_id, source_module, batch_label,
                status, workflow_status,
                total_debits, total_credits, summary_json
            ) VALUES (
                :entity_id, :accounting_period_id, :source_module, :batch_label,
                'draft', 'draft_ready',
                :total_debits, :total_credits, CAST(:summary_json AS jsonb)
            )
            ON CONFLICT (entity_id, accounting_period_id, source_module, batch_label)
            DO UPDATE SET
                status = 'draft',
                workflow_status = 'draft_ready',
                total_debits = EXCLUDED.total_debits,
                total_credits = EXCLUDED.total_credits,
                summary_json = EXCLUDED.summary_json,
                submitted_by = NULL, submitted_at = NULL,
                reviewed_by = NULL, reviewed_at = NULL,
                approved_by = NULL, approved_at = NULL,
                approval_note = NULL, rejection_note = NULL,
                locked_by = NULL, locked_at = NULL,
                updated_at = NOW()
            RETURNING id
            """
        ),
        {
            "entity_id": entity["id"],
            "accounting_period_id": accounting_period_id,
            "source_module": SOURCE_MODULE_DEPRECIATION,
            "batch_label": BATCH_LABEL_DEPRECIATION,
            "total_debits": total,
            "total_credits": total,
            "summary_json": json.dumps(summary),
        },
    ).mappings().first()
    journal_batch_id = batch["id"]

    # Wipe + rewrite journal_lines and depreciation_journal_lines for this batch.
    session.execute(
        text("DELETE FROM journal_lines WHERE journal_batch_id = :id"),
        {"id": journal_batch_id},
    )
    session.execute(
        text(
            """
            DELETE FROM depreciation_journal_lines
            WHERE entity_id = :entity_id AND period_start = :period_start
            """
        ),
        {"entity_id": entity["id"], "period_start": period_start},
    )

    for jl in journal_lines:
        session.execute(
            text(
                """
                INSERT INTO journal_lines (
                    journal_batch_id, line_number, account_code,
                    debit_amount, credit_amount, memo, source_json
                ) VALUES (
                    :id, :line_number, :account_code,
                    :debit_amount, :credit_amount, :memo, CAST(:src AS jsonb)
                )
                """
            ),
            {
                "id": journal_batch_id,
                "line_number": jl["line_number"],
                "account_code": jl["account_code"],
                "debit_amount": jl["debit_amount"],
                "credit_amount": jl["credit_amount"],
                "memo": jl["memo"],
                "src": json.dumps(
                    {
                        "source_module": SOURCE_MODULE_DEPRECIATION,
                        "fixed_asset_id": str(jl["asset_id"]),
                        "fiscal_year": fiscal_year,
                    }
                ),
            },
        )

    for r in rows:
        monthly = _money(r["monthly_depreciation"])
        if monthly == Decimal("0.00"):
            continue
        session.execute(
            text(
                """
                INSERT INTO depreciation_journal_lines (
                    entity_id, accounting_period_id, fixed_asset_id,
                    journal_batch_id,
                    period_start, period_end, monthly_depreciation,
                    debit_account, credit_account, posted_at
                ) VALUES (
                    :entity_id, :accounting_period_id, :asset_id,
                    :batch_id,
                    :period_start, :period_end, :monthly,
                    :dr, :cr, NOW()
                )
                """
            ),
            {
                "entity_id": entity["id"],
                "accounting_period_id": accounting_period_id,
                "asset_id": r["asset_id"],
                "batch_id": journal_batch_id,
                "period_start": period_start,
                "period_end": period_end,
                "monthly": monthly,
                "dr": r["depn_expense_gl_account"],
                "cr": r["accum_depn_gl_account"],
            },
        )

    return {
        "journal_batch_id": str(journal_batch_id),
        "entity_code": entity_code,
        "fiscal_year": fiscal_year,
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "asset_count": len(rows),
        "total_debits": str(total.quantize(Decimal("0.01"))),
        "total_credits": str(total.quantize(Decimal("0.01"))),
        "previously_posted_asset_count": existing_count,
    }


# ----------------------------------------------------------------------
# Reads
# ----------------------------------------------------------------------


def list_fixed_assets(session, *, entity_code: str) -> dict[str, Any]:
    entity = get_entity_by_code(session, entity_code)
    if not entity:
        raise ValueError(f"Unknown entity code: {entity_code}")
    if not _has_table(session, "fixed_assets"):
        return {"entity_code": entity_code, "count": 0, "assets": []}

    rows = session.execute(
        text(
            """
            SELECT id, asset_code, description, cca_class, cca_rate,
                   asset_gl_account, accum_depn_gl_account,
                   depn_expense_gl_account,
                   acquisition_date, cost, opening_nbv, opening_nbv_date,
                   is_active, disposal_date, disposal_proceeds, notes,
                   created_at
            FROM fixed_assets
            WHERE entity_id = :entity_id
            ORDER BY asset_code
            """
        ),
        {"entity_id": entity["id"]},
    ).mappings().all()

    return {
        "entity_code": entity_code,
        "count": len(rows),
        "assets": [
            {
                "id": str(r["id"]),
                "asset_code": r["asset_code"],
                "description": r["description"],
                "cca_class": r["cca_class"],
                "cca_rate": str(r["cca_rate"]),
                "asset_gl_account": r["asset_gl_account"],
                "accum_depn_gl_account": r["accum_depn_gl_account"],
                "depn_expense_gl_account": r["depn_expense_gl_account"],
                "acquisition_date": (
                    r["acquisition_date"].isoformat() if r["acquisition_date"] else None
                ),
                "cost": str(r["cost"]),
                "opening_nbv": str(r["opening_nbv"]),
                "opening_nbv_date": r["opening_nbv_date"].isoformat() if r["opening_nbv_date"] else None,
                "is_active": r["is_active"],
                "disposal_date": (
                    r["disposal_date"].isoformat() if r["disposal_date"] else None
                ),
                "disposal_proceeds": (
                    str(r["disposal_proceeds"]) if r["disposal_proceeds"] is not None else None
                ),
                "notes": r["notes"],
            }
            for r in rows
        ],
    }


def get_depreciation_schedule(
    session, *, entity_code: str, fiscal_year: int
) -> dict[str, Any]:
    entity = get_entity_by_code(session, entity_code)
    if not entity:
        raise ValueError(f"Unknown entity code: {entity_code}")

    rows = session.execute(
        text(
            """
            SELECT fa.asset_code, fa.description, fa.cca_class,
                   ds.opening_nbv, ds.annual_cca_rate, ds.half_year_rule_applies,
                   ds.annual_depreciation, ds.monthly_depreciation, ds.closing_nbv
            FROM depreciation_schedules ds
            JOIN fixed_assets fa ON fa.id = ds.fixed_asset_id
            WHERE ds.entity_id = :entity_id AND ds.fiscal_year = :fiscal_year
            ORDER BY fa.asset_code
            """
        ),
        {"entity_id": entity["id"], "fiscal_year": int(fiscal_year)},
    ).mappings().all()

    total_annual = sum(
        (Decimal(str(r["annual_depreciation"])) for r in rows), Decimal("0")
    )
    total_monthly = sum(
        (Decimal(str(r["monthly_depreciation"])) for r in rows), Decimal("0")
    )

    return {
        "entity_code": entity_code,
        "fiscal_year": int(fiscal_year),
        "row_count": len(rows),
        "total_annual_depreciation": str(total_annual.quantize(Decimal("0.01"))),
        "total_monthly_depreciation": str(total_monthly.quantize(Decimal("0.01"))),
        "rows": [
            {
                "asset_code": r["asset_code"],
                "description": r["description"],
                "cca_class": r["cca_class"],
                "opening_nbv": str(r["opening_nbv"]),
                "annual_cca_rate": str(r["annual_cca_rate"]),
                "half_year_rule_applies": r["half_year_rule_applies"],
                "annual_depreciation": str(r["annual_depreciation"]),
                "monthly_depreciation": str(r["monthly_depreciation"]),
                "closing_nbv": str(r["closing_nbv"]),
            }
            for r in rows
        ],
    }


def get_depreciation_summary(
    session, *, entity_code: str, period_end: date
) -> dict[str, Any]:
    entity = get_entity_by_code(session, entity_code)
    if not entity:
        raise ValueError(f"Unknown entity code: {entity_code}")

    fiscal_year = _fiscal_year_for_period(entity, period_end)

    monthly_rows = session.execute(
        text(
            """
            SELECT fa.cca_class, fa.description,
                   ds.monthly_depreciation, ds.opening_nbv, ds.closing_nbv
            FROM depreciation_schedules ds
            JOIN fixed_assets fa ON fa.id = ds.fixed_asset_id
            WHERE ds.entity_id = :entity_id AND ds.fiscal_year = :fiscal_year
            ORDER BY fa.cca_class
            """
        ),
        {"entity_id": entity["id"], "fiscal_year": fiscal_year},
    ).mappings().all()

    posted_rows = session.execute(
        text(
            """
            SELECT fa.cca_class,
                   COALESCE(SUM(djl.monthly_depreciation), 0) AS posted_total,
                   COUNT(*) AS posted_periods
            FROM depreciation_journal_lines djl
            JOIN fixed_assets fa ON fa.id = djl.fixed_asset_id
            WHERE djl.entity_id = :entity_id
              AND djl.period_end <= :period_end
              AND djl.period_start >= :ytd_start
            GROUP BY fa.cca_class
            """
        ),
        {
            "entity_id": entity["id"],
            "period_end": period_end,
            "ytd_start": date(
                fiscal_year - 1,
                int(entity.get("fiscal_year_end_month") or 9),
                int(entity.get("fiscal_year_end_day") or 30),
            ),
        },
    ).mappings().all()
    posted_by_class = {r["cca_class"]: r for r in posted_rows}

    by_class: list[dict[str, Any]] = []
    grand_monthly = Decimal("0")
    grand_ytd = Decimal("0")
    for r in monthly_rows:
        cls = r["cca_class"]
        monthly = Decimal(str(r["monthly_depreciation"]))
        ytd = Decimal(str(posted_by_class.get(cls, {}).get("posted_total") or 0))
        grand_monthly += monthly
        grand_ytd += ytd
        by_class.append(
            {
                "cca_class": cls,
                "description": r["description"],
                "monthly_depreciation": str(monthly),
                "ytd_posted": str(ytd),
                "current_nbv_estimate": str(
                    (Decimal(str(r["opening_nbv"])) - ytd).quantize(Decimal("0.01"))
                ),
            }
        )

    return {
        "entity_code": entity_code,
        "period_end": period_end.isoformat(),
        "fiscal_year": fiscal_year,
        "grand_total_monthly": str(grand_monthly.quantize(Decimal("0.01"))),
        "grand_total_ytd_posted": str(grand_ytd.quantize(Decimal("0.01"))),
        "by_class": by_class,
    }


# ======================================================================
# MODULE B ADDITIONS — asset-class CRUD, compute-by-class, Excel export,
# disposal journal, asset add.
# ======================================================================


# -----------------------------------------------------------------------
# fixed_asset_classes CRUD
# -----------------------------------------------------------------------

_BRIDLEWOOD_CLASSES = [
    {
        "class_code": CCA_CLASS_8_EQUIPMENT,
        "class_name": "Store Equipment",
        "cca_rate": Decimal("0.15"),
        "expense_account": DEPN_EXPENSE_GL_ACCOUNT,
        "accum_account": "1610",
        "display_order": 1,
    },
    {
        "class_code": CCA_CLASS_8_COMPUTER,
        "class_name": "Computer Equipment",
        "cca_rate": Decimal("0.15"),
        "expense_account": DEPN_EXPENSE_GL_ACCOUNT,
        "accum_account": "1640",
        "display_order": 2,
    },
    {
        "class_code": CCA_CLASS_10_VEHICLE,
        "class_name": "Vehicles",
        "cca_rate": Decimal("0.30"),
        "expense_account": DEPN_EXPENSE_GL_ACCOUNT,
        "accum_account": "1620",
        "display_order": 3,
    },
]


def seed_bridlewood_classes(
    session, *, entity_code: str
) -> dict[str, Any]:
    """Seed the 3 standard Bridlewood CCA classes into fixed_asset_classes.
    Idempotent — skips any class_code that already exists.
    Account codes verified against Bridlewood CoA (2026-06-06):
      1610 Acc. Depn Store Equipment
      1620 Acc. Depn Vehicles
      1640 Acc Depn Computers & System Software
      6900 Depreciation (expense)
    """
    entity = get_entity_by_code(session, entity_code)
    if not entity:
        raise ValueError(f"Unknown entity code: {entity_code}")

    inserted = 0
    skipped = 0
    ids: dict[str, str] = {}
    for cls in _BRIDLEWOOD_CLASSES:
        existing = session.execute(
            text(
                "SELECT id FROM fixed_asset_classes "
                "WHERE entity_id = :eid AND class_code = :cc"
            ),
            {"eid": entity["id"], "cc": cls["class_code"]},
        ).mappings().first()
        if existing:
            skipped += 1
            ids[cls["class_code"]] = str(existing["id"])
            continue
        row = session.execute(
            text(
                """
                INSERT INTO fixed_asset_classes (
                    entity_id, class_code, class_name, cca_rate,
                    expense_account, accum_account, display_order
                ) VALUES (
                    :eid, :cc, :cn, :rate, :exp_acct, :accum_acct, :order
                ) RETURNING id
                """
            ),
            {
                "eid": entity["id"],
                "cc": cls["class_code"],
                "cn": cls["class_name"],
                "rate": cls["cca_rate"],
                "exp_acct": cls["expense_account"],
                "accum_acct": cls["accum_account"],
                "order": cls["display_order"],
            },
        ).mappings().first()
        inserted += 1
        ids[cls["class_code"]] = str(row["id"])

    return {
        "entity_code": entity_code,
        "inserted": inserted,
        "skipped": skipped,
        "class_ids": ids,
    }


def link_assets_to_classes(
    session, *, entity_code: str
) -> dict[str, Any]:
    """Set fixed_asset_class_id on existing fixed_assets rows based on cca_class.
    MUST be called AFTER seed_bridlewood_classes.
    Safe to re-run (skips already-linked assets).
    """
    entity = get_entity_by_code(session, entity_code)
    if not entity:
        raise ValueError(f"Unknown entity code: {entity_code}")

    classes = session.execute(
        text(
            "SELECT id, class_code FROM fixed_asset_classes WHERE entity_id = :eid"
        ),
        {"eid": entity["id"]},
    ).mappings().all()
    class_map = {r["class_code"]: r["id"] for r in classes}
    if not class_map:
        return {"entity_code": entity_code, "linked": 0, "skipped": 0,
                "error": "No fixed_asset_classes found — run seed-classes first"}

    assets = session.execute(
        text(
            """SELECT id, asset_code, cca_class, fixed_asset_class_id
               FROM fixed_assets WHERE entity_id = :eid"""
        ),
        {"eid": entity["id"]},
    ).mappings().all()

    linked = 0
    skipped = 0
    not_found = []
    for a in assets:
        if a["fixed_asset_class_id"] is not None:
            skipped += 1
            continue
        class_id = class_map.get(a["cca_class"])
        if class_id is None:
            not_found.append(a["asset_code"])
            continue
        session.execute(
            text(
                "UPDATE fixed_assets SET fixed_asset_class_id = :cid WHERE id = :id"
            ),
            {"cid": class_id, "id": a["id"]},
        )
        linked += 1

    return {
        "entity_code": entity_code,
        "linked": linked,
        "skipped": skipped,
        "class_not_found_for": not_found,
    }


def list_asset_classes(session, *, entity_code: str) -> dict[str, Any]:
    entity = get_entity_by_code(session, entity_code)
    if not entity:
        raise ValueError(f"Unknown entity code: {entity_code}")

    rows = session.execute(
        text(
            """
            SELECT id, class_code, class_name, cca_rate, expense_account,
                   accum_account, formula_expr, is_active, display_order,
                   created_at, updated_at
            FROM fixed_asset_classes
            WHERE entity_id = :eid
            ORDER BY display_order, class_name
            """
        ),
        {"eid": entity["id"]},
    ).mappings().all()

    return {
        "entity_code": entity_code,
        "count": len(rows),
        "classes": [
            {
                "id": str(r["id"]),
                "class_code": r["class_code"],
                "class_name": r["class_name"],
                "cca_rate": str(r["cca_rate"]),
                "expense_account": r["expense_account"],
                "accum_account": r["accum_account"],
                "formula_expr": r["formula_expr"],
                "is_active": r["is_active"],
                "display_order": r["display_order"],
            }
            for r in rows
        ],
    }


def upsert_asset_class(
    session,
    *,
    entity_code: str,
    class_code: str,
    class_name: str,
    cca_rate: Decimal,
    expense_account: str,
    accum_account: str,
    formula_expr: str | None = None,
    is_active: bool = True,
    display_order: int = 0,
) -> dict[str, Any]:
    entity = get_entity_by_code(session, entity_code)
    if not entity:
        raise ValueError(f"Unknown entity code: {entity_code}")

    row = session.execute(
        text(
            """
            INSERT INTO fixed_asset_classes (
                entity_id, class_code, class_name, cca_rate,
                expense_account, accum_account, formula_expr,
                is_active, display_order
            ) VALUES (
                :eid, :cc, :cn, :rate, :exp, :accum, :formula,
                :active, :order
            )
            ON CONFLICT (entity_id, class_code) DO UPDATE SET
                class_name      = EXCLUDED.class_name,
                cca_rate        = EXCLUDED.cca_rate,
                expense_account = EXCLUDED.expense_account,
                accum_account   = EXCLUDED.accum_account,
                formula_expr    = EXCLUDED.formula_expr,
                is_active       = EXCLUDED.is_active,
                display_order   = EXCLUDED.display_order,
                updated_at      = NOW()
            RETURNING id
            """
        ),
        {
            "eid": entity["id"],
            "cc": class_code,
            "cn": class_name,
            "rate": cca_rate,
            "exp": expense_account,
            "accum": accum_account,
            "formula": formula_expr,
            "active": is_active,
            "order": display_order,
        },
    ).mappings().first()
    return {"id": str(row["id"]), "class_code": class_code}


# -----------------------------------------------------------------------
# Add fixed asset
# -----------------------------------------------------------------------


def add_fixed_asset(
    session,
    *,
    entity_code: str,
    asset_code: str,
    description: str,
    fixed_asset_class_id: str,
    acquisition_date: date,
    cost: Decimal,
    opening_nbv: Decimal | None = None,
    opening_nbv_date: date | None = None,
    notes: str | None = None,
    actor_email: str,
) -> dict[str, Any]:
    entity = get_entity_by_code(session, entity_code)
    if not entity:
        raise ValueError(f"Unknown entity code: {entity_code}")

    # Validate the class belongs to this entity and get account codes
    cls_row = session.execute(
        text(
            """
            SELECT id, class_code, cca_rate, expense_account, accum_account
            FROM fixed_asset_classes
            WHERE id = :cid AND entity_id = :eid
            """
        ),
        {"cid": fixed_asset_class_id, "eid": entity["id"]},
    ).mappings().first()
    if not cls_row:
        raise ValueError(f"fixed_asset_class_id not found for entity: {fixed_asset_class_id}")

    row = session.execute(
        text(
            """
            INSERT INTO fixed_assets (
                entity_id, asset_code, description,
                cca_class, cca_rate,
                asset_gl_account, accum_depn_gl_account, depn_expense_gl_account,
                fixed_asset_class_id, acquisition_date, cost,
                opening_nbv, opening_nbv_date, is_active, notes
            ) VALUES (
                :eid, :ac, :desc,
                :cc, :rate,
                :asset_gl, :accum_gl, :exp_gl,
                :class_id, :acq_date, :cost,
                :onbv, :onbv_date, TRUE, :notes
            ) RETURNING id
            """
        ),
        {
            "eid": entity["id"],
            "ac": asset_code,
            "desc": description,
            "cc": cls_row["class_code"],
            "rate": cls_row["cca_rate"],
            "asset_gl": "1500",          # caller can override; generic asset account
            "accum_gl": cls_row["accum_account"],
            "exp_gl": cls_row["expense_account"],
            "class_id": fixed_asset_class_id,
            "acq_date": acquisition_date,
            "cost": cost,
            "onbv": opening_nbv if opening_nbv is not None else cost,
            "onbv_date": opening_nbv_date or acquisition_date,
            "notes": notes or f"Added by {actor_email}",
        },
    ).mappings().first()
    return {"id": str(row["id"]), "asset_code": asset_code}


# -----------------------------------------------------------------------
# compute_monthly_depreciation_by_class  (Module A interface)
# -----------------------------------------------------------------------


def compute_monthly_depreciation_by_class(
    session,
    *,
    entity_id: UUID,
    period_end: date,
) -> list[dict[str, Any]]:
    """Return per-class monthly depreciation amounts for a given period.

    Used by Module A (recurring entry engine) as the SCHEDULE feeder.
    Pulls from depreciation_schedules joined to fixed_asset_classes via the
    fixed_asset_class_id FK.  Falls back to the cca_class TEXT field for
    assets not yet linked to a class row.

    Returns a list of dicts:
        {class_id, class_code, class_name, expense_account, accum_account, amount}
    Empty list if no schedule data exists.
    """
    # Fiscal year for the requested period (needed to look up the schedule)
    entity_row = session.execute(
        text("SELECT fiscal_year_end_month, fiscal_year_end_day FROM entities WHERE id = :eid"),
        {"eid": entity_id},
    ).mappings().first()
    fy_month = int(entity_row["fiscal_year_end_month"] or 9) if entity_row else 9
    fy_day = int(entity_row["fiscal_year_end_day"] or 30) if entity_row else 30
    if (period_end.month, period_end.day) <= (fy_month, fy_day):
        fiscal_year = period_end.year
    else:
        fiscal_year = period_end.year + 1

    # Preferred path: assets with fixed_asset_class_id set
    rows = session.execute(
        text(
            """
            SELECT
                fac.id           AS class_id,
                fac.class_code,
                fac.class_name,
                fac.expense_account,
                fac.accum_account,
                SUM(ds.monthly_depreciation) AS amount
            FROM depreciation_schedules ds
            JOIN fixed_assets fa
                ON fa.id = ds.fixed_asset_id AND fa.entity_id = ds.entity_id
            JOIN fixed_asset_classes fac
                ON fac.id = fa.fixed_asset_class_id
            WHERE ds.entity_id = :eid
              AND ds.fiscal_year = :fy
              AND fa.is_active = TRUE
              AND (fa.disposal_date IS NULL OR fa.disposal_date >= :period_end)
            GROUP BY fac.id, fac.class_code, fac.class_name,
                     fac.expense_account, fac.accum_account
            ORDER BY fac.display_order, fac.class_name
            """
        ),
        {"eid": entity_id, "fy": fiscal_year, "period_end": period_end},
    ).mappings().all()

    if rows:
        return [
            {
                "class_id": str(r["class_id"]),
                "class_code": r["class_code"],
                "class_name": r["class_name"],
                "expense_account": r["expense_account"],
                "accum_account": r["accum_account"],
                "amount": _money(r["amount"]),
            }
            for r in rows
        ]

    # Fallback: no fixed_asset_class_id set — group by the text cca_class field
    fallback = session.execute(
        text(
            """
            SELECT
                fa.cca_class,
                fa.depn_expense_gl_account   AS expense_account,
                fa.accum_depn_gl_account     AS accum_account,
                SUM(ds.monthly_depreciation) AS amount
            FROM depreciation_schedules ds
            JOIN fixed_assets fa
                ON fa.id = ds.fixed_asset_id AND fa.entity_id = ds.entity_id
            WHERE ds.entity_id = :eid
              AND ds.fiscal_year = :fy
              AND fa.is_active = TRUE
              AND (fa.disposal_date IS NULL OR fa.disposal_date >= :period_end)
            GROUP BY fa.cca_class, fa.depn_expense_gl_account, fa.accum_depn_gl_account
            ORDER BY fa.cca_class
            """
        ),
        {"eid": entity_id, "fy": fiscal_year, "period_end": period_end},
    ).mappings().all()

    return [
        {
            "class_id": None,
            "class_code": r["cca_class"],
            "class_name": r["cca_class"],
            "expense_account": r["expense_account"],
            "accum_account": r["accum_account"],
            "amount": _money(r["amount"]),
        }
        for r in fallback
    ]


# -----------------------------------------------------------------------
# Disposal journal  (NOTE: gain/loss accounts TBD for Bridlewood)
# -----------------------------------------------------------------------


class MissingGainLossAccountError(ValueError):
    """Raised when the disposal gain/loss GL accounts are not configured."""


def compute_disposal_nbv(
    session, *, entity_id: UUID, fixed_asset_id: str, as_of: date
) -> Decimal:
    """Compute the NBV of an asset at a given date (opening_nbv minus YTD depreciation posted)."""
    asset = session.execute(
        text(
            "SELECT opening_nbv, opening_nbv_date FROM fixed_assets "
            "WHERE id = :id AND entity_id = :eid"
        ),
        {"id": fixed_asset_id, "eid": entity_id},
    ).mappings().first()
    if not asset:
        raise ValueError(f"fixed_asset not found: {fixed_asset_id}")

    ytd_posted = session.execute(
        text(
            """
            SELECT COALESCE(SUM(monthly_depreciation), 0) AS total
            FROM depreciation_journal_lines
            WHERE entity_id = :eid
              AND fixed_asset_id = :asset_id
              AND period_end <= :as_of
              AND period_end >= :from_date
            """
        ),
        {
            "eid": entity_id,
            "asset_id": fixed_asset_id,
            "as_of": as_of,
            "from_date": asset["opening_nbv_date"],
        },
    ).mappings().first()
    total_posted = _money(ytd_posted["total"] or 0)
    return _money(asset["opening_nbv"]) - total_posted


def post_disposal_journal(
    session,
    *,
    entity_code: str,
    fixed_asset_id: str,
    disposal_date: date,
    proceeds: Decimal,
    proceeds_account: str,
    gain_account: str,
    loss_account: str,
    actor_email: str,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Post the disposal journal entry for a fixed asset.

    IMPORTANT — Bridlewood CoA as of 2026-06-06:
        gain_account (4020) and loss_account (6950) do NOT exist in the
        chart of accounts. Call this with verified accounts only.
        Raise MissingGainLossAccountError if either is not supplied.

    Journal (gain scenario):
        Dr  accum_depn_gl_account   (accumulated depreciation removed)
        Dr  proceeds_account        (cash/AR received)
        Cr  asset_gl_account        (cost removed)
        Cr  gain_account            (gain recognised)

    Journal (loss scenario):
        Dr  accum_depn_gl_account
        Dr  proceeds_account
        Dr  loss_account            (loss recognised)
        Cr  asset_gl_account
    """
    from .services_period_close import assert_period_not_locked

    if not gain_account or not loss_account:
        raise MissingGainLossAccountError(
            "gain_account and loss_account must be supplied and exist in the "
            "chart of accounts. Bridlewood does not have 4020/6950 as of "
            "the last CoA check — confirm with the bookkeeper before proceeding."
        )

    entity = get_entity_by_code(session, entity_code)
    if not entity:
        raise ValueError(f"Unknown entity code: {entity_code}")

    asset = session.execute(
        text(
            """
            SELECT id, asset_code, description, asset_gl_account,
                   accum_depn_gl_account, depn_expense_gl_account, cost,
                   fixed_asset_class_id
            FROM fixed_assets
            WHERE id = :id AND entity_id = :eid AND is_active = TRUE
            """
        ),
        {"id": fixed_asset_id, "eid": entity["id"]},
    ).mappings().first()
    if not asset:
        raise ValueError(f"Active fixed_asset not found: {fixed_asset_id}")

    assert_period_not_locked(session, entity["id"], disposal_date)

    nbv = compute_disposal_nbv(
        session, entity_id=entity["id"],
        fixed_asset_id=fixed_asset_id, as_of=disposal_date
    )
    proceeds = _money(proceeds)
    gain_loss = _money(proceeds - nbv)  # positive = gain, negative = loss

    # Determine the total accumulated depreciation posted to date
    accum_depr = _money(
        _money(
            session.execute(
                text(
                    """SELECT COALESCE(SUM(monthly_depreciation), 0)
                       FROM depreciation_journal_lines
                       WHERE entity_id = :eid AND fixed_asset_id = :aid"""
                ),
                {"eid": entity["id"], "aid": fixed_asset_id},
            ).scalar() or 0
        )
    )

    # Build lines
    cost = _money(asset["cost"])
    lines: list[dict[str, Any]] = []
    line_num = 0

    def _add(acct: str, dr: Decimal, cr: Decimal, memo: str) -> None:
        nonlocal line_num
        line_num += 1
        lines.append(
            {
                "line_number": line_num,
                "account_code": acct,
                "debit_amount": dr,
                "credit_amount": cr,
                "memo": memo,
            }
        )

    _add(asset["accum_depn_gl_account"], accum_depr, Decimal("0"),
         f"Disposal of {asset['description']} — remove accumulated depreciation")
    if proceeds > Decimal("0"):
        _add(proceeds_account, proceeds, Decimal("0"),
             f"Disposal proceeds — {asset['description']}")
    _add(asset["asset_gl_account"], Decimal("0"), cost,
         f"Disposal of {asset['description']} — remove asset cost")
    if gain_loss > Decimal("0"):
        _add(gain_account, Decimal("0"), gain_loss,
             f"Gain on disposal — {asset['description']}")
    elif gain_loss < Decimal("0"):
        _add(loss_account, -gain_loss, Decimal("0"),
             f"Loss on disposal — {asset['description']}")

    total_dr = sum(l["debit_amount"] for l in lines)
    total_cr = sum(l["credit_amount"] for l in lines)
    if abs(total_dr - total_cr) > Decimal("0.01"):
        raise ValueError(
            f"Disposal journal is unbalanced: Dr={total_dr} Cr={total_cr}"
        )

    preview = {
        "asset_code": asset["asset_code"],
        "description": asset["description"],
        "disposal_date": disposal_date.isoformat(),
        "proceeds": str(proceeds),
        "nbv_at_disposal": str(nbv),
        "gain_loss": str(gain_loss),
        "accumulated_depreciation_removed": str(accum_depr),
        "journal_lines": [
            {
                "line_number": l["line_number"],
                "account_code": l["account_code"],
                "debit_amount": str(l["debit_amount"]),
                "credit_amount": str(l["credit_amount"]),
                "memo": l["memo"],
            }
            for l in lines
        ],
        "total_debits": str(total_dr),
        "total_credits": str(total_cr),
    }

    if dry_run:
        return {"dry_run": True, **preview}

    # Post the journal batch
    from .services import get_or_create_accounting_period
    accounting_period_id = get_or_create_accounting_period(
        session, entity["id"], disposal_date
    )
    batch = session.execute(
        text(
            """
            INSERT INTO journal_batches (
                entity_id, accounting_period_id, source_module, batch_label,
                status, workflow_status, total_debits, total_credits, summary_json
            ) VALUES (
                :eid, :apid, 'fixed_asset_disposal',
                :label, 'posted', 'posted',
                :dr, :cr, CAST(:sj AS jsonb)
            ) RETURNING id
            """
        ),
        {
            "eid": entity["id"],
            "apid": accounting_period_id,
            "label": f"disposal_{asset['asset_code']}_{disposal_date.isoformat()}",
            "dr": total_dr,
            "cr": total_cr,
            "sj": json.dumps(preview),
        },
    ).mappings().first()
    batch_id = batch["id"]

    for ln in lines:
        session.execute(
            text(
                """
                INSERT INTO journal_lines (
                    journal_batch_id, line_number, account_code,
                    debit_amount, credit_amount, memo
                ) VALUES (:bid, :ln, :ac, :dr, :cr, :memo)
                """
            ),
            {
                "bid": batch_id,
                "ln": ln["line_number"],
                "ac": ln["account_code"],
                "dr": ln["debit_amount"],
                "cr": ln["credit_amount"],
                "memo": ln["memo"],
            },
        )

    # Record the disposal
    session.execute(
        text(
            """
            INSERT INTO fixed_asset_disposals (
                entity_id, fixed_asset_id, disposal_date, proceeds,
                nbv_at_disposal, gain_loss, proceeds_account,
                gain_loss_account, journal_batch_id
            ) VALUES (
                :eid, :aid, :dd, :proc,
                :nbv, :gl, :pa, :gla, :bid
            )
            """
        ),
        {
            "eid": entity["id"],
            "aid": fixed_asset_id,
            "dd": disposal_date,
            "proc": proceeds,
            "nbv": nbv,
            "gl": gain_loss,
            "pa": proceeds_account,
            "gla": gain_account if gain_loss >= 0 else loss_account,
            "bid": batch_id,
        },
    )

    # Mark asset as disposed
    session.execute(
        text(
            """
            UPDATE fixed_assets
            SET is_active = FALSE, disposal_date = :dd, disposal_proceeds = :proc
            WHERE id = :id
            """
        ),
        {"dd": disposal_date, "proc": proceeds, "id": fixed_asset_id},
    )

    return {"dry_run": False, "journal_batch_id": str(batch_id), **preview}


# -----------------------------------------------------------------------
# Excel schedule export
# -----------------------------------------------------------------------


def generate_excel_schedule(
    session,
    *,
    entity_id: UUID,
    fiscal_year: int,
) -> bytes:
    """Generate an openpyxl workbook (bytes) matching the Bridlewood
    fixed-asset schedule format.

    Columns (16, in order):
      1  Date of Addition
      2  Description
      3  Cost — Opening Balance
      4  Additions
      5  Disposals
      6  Cost — Closing Balance
      7  Salvage Value   (always blank for CCA)
      8  Proceeds on Disposal
      9  Gain / (Loss)
      10 Method / Life (CCA rate as %)
      11 Accumulated Depr — Opening Balance
      12 NBV Before Depreciation
      13 Depreciation
      14 Dispositions / Other
      15 Accumulated Depr — Closing Balance
      16 NBV — Closing

    Grouped by asset class with per-class subtotals + grand total.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError("openpyxl is required for Excel schedule export") from exc

    # Pull asset data and schedules
    rows = session.execute(
        text(
            """
            SELECT
                fa.id,
                fa.asset_code,
                fa.description,
                fa.acquisition_date,
                fa.cost,
                fa.opening_nbv,
                fac.class_name,
                fac.cca_rate,
                fac.display_order  AS class_order,
                ds.opening_nbv     AS sched_opening_nbv,
                ds.annual_depreciation,
                ds.monthly_depreciation,
                ds.closing_nbv     AS sched_closing_nbv,
                -- Disposals this FY
                COALESCE(fad.proceeds, 0)    AS disposal_proceeds,
                COALESCE(fad.gain_loss, 0)   AS disposal_gain_loss,
                fad.disposal_date
            FROM fixed_assets fa
            LEFT JOIN fixed_asset_classes fac
                ON fac.id = fa.fixed_asset_class_id
            LEFT JOIN depreciation_schedules ds
                ON ds.fixed_asset_id = fa.id
                   AND ds.entity_id = fa.entity_id
                   AND ds.fiscal_year = :fy
            LEFT JOIN fixed_asset_disposals fad
                ON fad.fixed_asset_id = fa.id
                   AND EXTRACT(YEAR FROM fad.disposal_date) = :fy
            WHERE fa.entity_id = :eid
            ORDER BY COALESCE(fac.display_order, 999), fac.class_name, fa.asset_code
            """
        ),
        {"eid": entity_id, "fy": fiscal_year},
    ).mappings().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Fixed Assets FY{fiscal_year}"

    # Styles
    header_font = Font(bold=True, size=10)
    class_font = Font(bold=True, size=10, italic=True)
    subtotal_font = Font(bold=True, size=10)
    currency_fmt = '#,##0.00'
    pct_fmt = '0.00%'

    HEADERS = [
        "Date of Addition", "Description",
        "Cost\nOpening Balance", "Additions", "Disposals", "Cost\nClosing Balance",
        "Salvage\nValue", "Proceeds on\nDisposal", "Gain / (Loss)",
        "Method/Life\n(CCA Rate)",
        "Accum. Depr.\nOpening Balance", "NBV Before\nDepreciation",
        "Depreciation", "Dispositions/\nOther",
        "Accum. Depr.\nClosing Balance", "NBV\nClosing",
    ]

    # Title row
    ws.merge_cells("A1:P1")
    title_cell = ws["A1"]
    title_cell.value = f"Bridlewood Hardware Co. Ltd. — Fixed Assets Schedule FY{fiscal_year}"
    title_cell.font = Font(bold=True, size=12)
    ws.row_dimensions[1].height = 20

    # Header row
    for col, hdr in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
    ws.row_dimensions[2].height = 32

    # Column widths
    col_widths = [14, 32, 14, 12, 12, 14, 10, 12, 12, 12, 16, 14, 12, 12, 16, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    current_row = 3

    def _v(val: Any) -> Decimal:
        return _money(val) if val is not None else Decimal("0")

    # Group by class
    from itertools import groupby
    key_fn = lambda r: (r["class_order"] or 999, r["class_name"] or "Uncategorised")
    groups = groupby(rows, key=key_fn)

    grand = {k: Decimal("0") for k in
             ["cost_open", "additions", "disposals", "cost_close",
              "proceeds", "gain_loss", "accum_open", "depr", "accum_close"]}

    for _, asset_rows in groups:
        asset_list = list(asset_rows)
        if not asset_list:
            continue

        class_name = asset_list[0]["class_name"] or "Uncategorised"
        cca_rate = _v(asset_list[0]["cca_rate"])

        # Class header
        ws.cell(row=current_row, column=1, value=class_name).font = class_font
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=len(HEADERS)
        )
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        sub = {k: Decimal("0") for k in grand}

        for r in asset_list:
            cost_open = _v(r["cost"])
            additions = Decimal("0")      # future: track via additions table
            disposals = _v(r["disposal_proceeds"])
            cost_close = cost_open + additions - disposals
            proceeds = _v(r["disposal_proceeds"])
            gain_loss = _v(r["disposal_gain_loss"])
            accum_open = _v(r["opening_nbv"]) - _v(r["cost"])  # opening accum = cost - opening_nbv
            # If schedule row exists
            depr = _v(r["annual_depreciation"])
            nbv_before = _v(r["sched_opening_nbv"]) if r["sched_opening_nbv"] else _v(r["opening_nbv"])
            accum_close = accum_open + depr
            nbv_close = _v(r["sched_closing_nbv"]) if r["sched_closing_nbv"] else (nbv_before - depr)

            data = [
                r["acquisition_date"],
                r["description"],
                float(cost_open),
                float(additions),
                float(disposals),
                float(cost_close),
                None,              # salvage value — always blank for CCA
                float(proceeds) if proceeds else None,
                float(gain_loss) if gain_loss else None,
                float(cca_rate),
                float(accum_open),
                float(nbv_before),
                float(depr),
                None,              # dispositions/other
                float(accum_close),
                float(nbv_close),
            ]
            for col, val in enumerate(data, start=1):
                cell = ws.cell(row=current_row, column=col, value=val)
                if col == 10:
                    cell.number_format = pct_fmt
                elif col > 2 and val is not None:
                    cell.number_format = currency_fmt
            current_row += 1

            sub["cost_open"] += cost_open
            sub["additions"] += additions
            sub["disposals"] += disposals
            sub["cost_close"] += cost_close
            sub["proceeds"] += proceeds
            sub["gain_loss"] += gain_loss
            sub["accum_open"] += accum_open
            sub["depr"] += depr
            sub["accum_close"] += accum_close

        # Subtotal row
        sub_data = [
            None, f"{class_name} Subtotal",
            float(sub["cost_open"]), float(sub["additions"]),
            float(sub["disposals"]), float(sub["cost_close"]),
            None, float(sub["proceeds"]), float(sub["gain_loss"]),
            None,
            float(sub["accum_open"]), None,
            float(sub["depr"]), None, float(sub["accum_close"]), None,
        ]
        for col, val in enumerate(sub_data, start=1):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.font = subtotal_font
            if col > 2 and val is not None:
                cell.number_format = currency_fmt
        current_row += 2

        for k in grand:
            grand[k] += sub[k]

    # Grand total
    grand_data = [
        None, "GRAND TOTAL",
        float(grand["cost_open"]), float(grand["additions"]),
        float(grand["disposals"]), float(grand["cost_close"]),
        None, float(grand["proceeds"]), float(grand["gain_loss"]),
        None,
        float(grand["accum_open"]), None,
        float(grand["depr"]), None, float(grand["accum_close"]), None,
    ]
    for col, val in enumerate(grand_data, start=1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = Font(bold=True, size=11)
        if col > 2 and val is not None:
            cell.number_format = currency_fmt

    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------------------------------------------------------
# Close-control-center section
# ----------------------------------------------------------------------


def section_depreciation(
    session,
    *,
    entity_id: UUID,
    accounting_period_id: UUID,
    period_end: date,
) -> dict[str, Any]:
    if not _has_table(session, "depreciation_journal_lines"):
        return {
            "status": "no_data",
            "module_present": False,
            "summary": "depreciation_journal_lines table not present",
        }

    period_start = period_end.replace(day=1)
    posted = session.execute(
        text(
            """
            SELECT fixed_asset_id, monthly_depreciation, journal_batch_id
            FROM depreciation_journal_lines
            WHERE entity_id = :entity_id
              AND period_start = :period_start
            """
        ),
        {"entity_id": entity_id, "period_start": period_start},
    ).mappings().all()

    active_count = session.execute(
        text(
            """
            SELECT COUNT(*) AS n FROM fixed_assets
            WHERE entity_id = :entity_id AND is_active = TRUE
            """
        ),
        {"entity_id": entity_id},
    ).mappings().first()["n"] or 0

    if not posted:
        if active_count == 0:
            return {
                "status": "needs_review",
                "module_present": True,
                "summary": (
                    "No fixed assets seeded. POST /api/depreciation/seed-assets."
                ),
            }
        return {
            "status": "blocked",
            "module_present": True,
            "summary": (
                f"No depreciation journal posted for "
                f"{period_end.isoformat()} (active assets: {active_count}). "
                "POST /api/depreciation/build-journal."
            ),
        }

    total = sum(
        (Decimal(str(r["monthly_depreciation"])) for r in posted), Decimal("0")
    )
    return {
        "status": "ready",
        "module_present": True,
        "summary": (
            f"Depreciation posted for {len(posted)} of {active_count} assets, "
            f"total ${total.quantize(Decimal('0.01'))}."
        ),
        "asset_count_posted": len(posted),
        "active_asset_count": active_count,
        "total_posted": str(total.quantize(Decimal("0.01"))),
    }
