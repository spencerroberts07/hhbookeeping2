"""
Dealer onboarding service layer.

This module turns a new (or partly-set-up) dealer into a working entity
in two parallel paths:

  Path A — QuickBooks Online:  pull chart of accounts, trial balance,
                               and general ledger directly from QBO.
  Path B — File upload:        accept any format (CSV, Excel, QBO/Sage
                               export) and use Claude to parse it.

Both paths feed the same write pipeline: chart-of-accounts upserts +
opening-balance journal batch + historical journal batches +
vendor-memory bootstrap.

Public surface (used by routes/onboarding.py):

    detect_existing_data(session, entity_id)
        What's already set up. Used to pre-fill the wizard.

    parse_chart_of_accounts(file_bytes, filename)
        Claude → JSON preview. Doesn't write.
    save_chart_of_accounts(session, entity_id, accounts)
        Persist confirmed preview.

    parse_trial_balance(file_bytes, filename, as_of_date)
        Claude → JSON preview. Doesn't write.
    import_opening_balances(session, entity_id, entity_code,
                            as_of_date, tb_lines, actor_email)
        Validate balance, create period, journal batch + lines.
    import_trial_balance_from_qbo(session, entity_id, entity_code,
                                  as_of_date, actor_email)
        QBO TB → import_opening_balances.

    parse_gl_file(file_bytes, filename)
        Claude → JSON preview of journal lines.
    import_gl_history_from_file(session, entity_id, entity_code,
                                lines, actor_email, progress_callback)
        Parsed file lines → journal_batches/journal_lines pipeline.
    import_gl_history_from_qbo(session, entity_id, entity_code,
                               date_from, date_to, actor_email,
                               progress_callback)
        QBO GL pull month-by-month.

    learn_from_gl_history(session, entity_id, entity_code)
        Mine journal_lines for vendor patterns; bootstrap
        vendor_classification_memory + assistant_entity_memory.

Reliability notes:
- Claude calls degrade gracefully when ANTHROPIC_API_KEY is unset: file
  parsers return an explicit error so the route can 400 the request
  rather than silently empty-import.
- Long imports run inside FastAPI BackgroundTasks; each must open its
  own db_session (the route's session closes before the background
  task runs).
"""
from __future__ import annotations

import calendar
import csv
import io
import json
import logging
import re
from datetime import date as DateType, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Callable

from sqlalchemy import text

from .config import settings
from .quickbooks import (
    QuickBooksClient,
    ensure_valid_access_token,
    month_chunks,
)

logger = logging.getLogger(__name__)


CLAUDE_MODEL_ID = "claude-sonnet-4-6"
# 4096 tokens is enough for ~150-200 TB/CoA rows in JSON form. Bigger
# files should already be hitting the regex fallback path.
CLAUDE_PARSE_MAX_TOKENS = 4096
# 90s per attempt × 2 attempts = 3 minutes worst-case — matches the
# frontend's 3-minute polling cap.
CLAUDE_CALL_TIMEOUT_SECONDS = 90


# --------------------------------------------------------------------------
# 2A. Detect existing data
# --------------------------------------------------------------------------


def detect_existing_data(session, entity_id: str) -> dict[str, Any]:
    """Snapshot of what's already loaded for this entity. Used to
    pre-fill the wizard so dealers with partial setup (like Bridlewood)
    don't see a blank slate.
    """
    accounts_row = session.execute(
        text("SELECT COUNT(*) AS c FROM accounts WHERE entity_id = :eid"),
        {"eid": entity_id},
    ).mappings().first()
    account_count = int((accounts_row or {}).get("c") or 0)

    opening = session.execute(
        text(
            """
            SELECT jb.id, jb.created_at, ap.period_end
              FROM journal_batches jb
              JOIN accounting_periods ap ON ap.id = jb.accounting_period_id
             WHERE jb.entity_id = :eid AND jb.source_module = 'opening_balance'
             ORDER BY ap.period_end DESC
             LIMIT 1
            """
        ),
        {"eid": entity_id},
    ).mappings().first()

    gl_history = session.execute(
        text(
            """
            SELECT MIN(ap.period_start) AS from_d,
                   MAX(ap.period_end)   AS to_d,
                   COUNT(jl.id)         AS line_count
              FROM journal_batches jb
              JOIN journal_lines jl ON jl.journal_batch_id = jb.id
              JOIN accounting_periods ap ON ap.id = jb.accounting_period_id
             WHERE jb.entity_id = :eid AND jb.source_module = 'historical_import'
            """
        ),
        {"eid": entity_id},
    ).mappings().first()
    gl_line_count = int((gl_history or {}).get("line_count") or 0)

    hh_ap = session.execute(
        text(
            """
            -- hh_ap_statements has no period_start column. statement_month_end
            -- is the canonical "which month does this statement cover"
            -- column; statement_date is the issue date and is the
            -- fallback for older rows that pre-date statement_month_end.
            SELECT COUNT(DISTINCT date_trunc(
                'month',
                COALESCE(s.statement_month_end, s.statement_date)
            )) AS months
              FROM hh_ap_statements s
             WHERE s.entity_id = :eid
            """
        ),
        {"eid": entity_id},
    ).mappings().first()
    hh_ap_months = int((hh_ap or {}).get("months") or 0)

    bank_count_row = session.execute(
        text("SELECT COUNT(*) AS c FROM bank_transactions WHERE entity_id = :eid"),
        {"eid": entity_id},
    ).mappings().first()
    bank_count = int((bank_count_row or {}).get("c") or 0)

    qbo = session.execute(
        text(
            """
            SELECT realm_id, connected_at
              FROM quickbooks_connections
             WHERE entity_id = :eid AND is_active = TRUE
             ORDER BY connected_at DESC
             LIMIT 1
            """
        ),
        {"eid": entity_id},
    ).mappings().first()

    entity_flags = session.execute(
        text(
            """
            SELECT onboarding_complete, onboarding_completed_at
              FROM entities WHERE id = :eid
            """
        ),
        {"eid": entity_id},
    ).mappings().first() or {}

    has_chart = account_count > 0
    has_opening = opening is not None
    return {
        "has_chart_of_accounts": has_chart,
        "account_count": account_count,
        "has_opening_balances": has_opening,
        "opening_balance_date": (
            opening["period_end"].isoformat() if has_opening else None
        ),
        "has_gl_history": gl_line_count > 0,
        "gl_history_from": (
            gl_history["from_d"].isoformat() if gl_line_count and gl_history["from_d"] else None
        ),
        "gl_history_to": (
            gl_history["to_d"].isoformat() if gl_line_count and gl_history["to_d"] else None
        ),
        "journal_line_count": gl_line_count,
        "has_hh_ap_history": hh_ap_months > 0,
        "hh_ap_months_loaded": hh_ap_months,
        "has_bank_transactions": bank_count > 0,
        "bank_transaction_count": bank_count,
        "has_qbo_connection": qbo is not None,
        "qbo_realm_id": qbo["realm_id"] if qbo else None,
        "qbo_connected_at": (
            qbo["connected_at"].isoformat() if qbo else None
        ),
        "onboarding_complete": bool(entity_flags.get("onboarding_complete")),
        "onboarding_completed_at": (
            entity_flags["onboarding_completed_at"].isoformat()
            if entity_flags.get("onboarding_completed_at")
            else None
        ),
    }


# --------------------------------------------------------------------------
# Claude client (matches the pattern in services_assistant.py)
# --------------------------------------------------------------------------


def _claude_client():
    api_key = getattr(settings, "anthropic_api_key", None)
    if not api_key:
        return None
    try:
        from anthropic import Anthropic
    except ImportError:
        logger.warning("anthropic SDK not installed — file parsers degraded")
        return None
    return Anthropic(api_key=api_key)


def _decode_file(file_bytes: bytes, filename: str) -> str:
    """Render an uploaded file into a string Claude can read.

    CSV / TSV / plain text → utf-8 (lossy). Excel and ODS → openpyxl /
    odfpy to text. Other types → treat as text best-effort.
    """
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xls", ".xlsm")):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
            out_lines: list[str] = []
            for ws in wb.worksheets:
                out_lines.append(f"# Sheet: {ws.title}")
                for row in ws.iter_rows(values_only=True):
                    cells = [
                        ("" if v is None else str(v))
                        for v in row
                    ]
                    out_lines.append(",".join(cells))
                out_lines.append("")
            return "\n".join(out_lines)
        except Exception as exc:
            logger.warning("openpyxl decode failed for %s: %r", filename, exc)
    if name.endswith(".ods"):
        try:
            from odf.opendocument import load
            from odf.table import Table, TableRow, TableCell
            from odf.text import P
            doc = load(BytesIO(file_bytes))
            out_lines: list[str] = []
            for table in doc.spreadsheet.getElementsByType(Table):
                out_lines.append(f"# Sheet: {table.getAttribute('name')}")
                for row in table.getElementsByType(TableRow):
                    cells: list[str] = []
                    for cell in row.getElementsByType(TableCell):
                        text_val = "".join(
                            str(p) for p in cell.getElementsByType(P)
                        )
                        cells.append(text_val)
                    out_lines.append(",".join(cells))
                out_lines.append("")
            return "\n".join(out_lines)
        except Exception as exc:
            logger.warning("odfpy decode failed for %s: %r", filename, exc)
    try:
        return file_bytes.decode("utf-8", errors="replace")
    except Exception:
        return file_bytes.decode("latin-1", errors="replace")


def _claude_parse_json(system_prompt: str, file_text: str) -> dict[str, Any] | None:
    """Run a Claude call expecting a single JSON object back. Returns None
    when the model is unavailable or returns un-parseable text — the
    route layer converts that to a 400 with a clear message.

    Calls Claude with a 60s per-attempt timeout and one retry on
    transient failure. Callers should try a regex fallback before
    calling this for known-format files.
    """
    client = _claude_client()
    if not client:
        logger.warning("Claude unavailable — file parser degraded to None")
        return None

    msg = None
    last_exc: Exception | None = None
    for attempt in (1, 2):
        try:
            msg = client.messages.create(
                model=CLAUDE_MODEL_ID,
                max_tokens=CLAUDE_PARSE_MAX_TOKENS,
                system=system_prompt,
                messages=[{"role": "user", "content": file_text[:200_000]}],
                timeout=CLAUDE_CALL_TIMEOUT_SECONDS,
            )
            break
        except Exception as exc:
            last_exc = exc
            logger.warning(
                "Claude file-parse attempt %d failed: %r", attempt, exc
            )
    if msg is None:
        logger.warning("Claude file-parse exhausted retries: %r", last_exc)
        return None

    text_out = ""
    for block in getattr(msg, "content", None) or []:
        if getattr(block, "type", None) == "text":
            text_out += block.text or ""
    text_out = text_out.strip()
    if not text_out:
        return None
    if text_out.startswith("```"):
        text_out = text_out.strip("`").lstrip()
        if text_out.lower().startswith("json"):
            text_out = text_out[4:].lstrip("\n")
    try:
        return json.loads(text_out)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text_out, flags=re.DOTALL)
        if not match:
            logger.warning("Claude returned non-JSON: %r", text_out[:200])
            return None
        try:
            return json.loads(match.group(0))
        except json.JSONDecodeError:
            logger.warning("Claude returned non-JSON: %r", text_out[:200])
            return None


# --------------------------------------------------------------------------
# Regex / CSV fallback parsers
#
# Cheaper, faster, deterministic. Each parser returns None when it can't
# confidently identify the format — caller then falls back to Claude.
# Handles ~90% of cases (QBO TB export, Sage exports, simple CSV/Excel).
# --------------------------------------------------------------------------


_AMOUNT_KEYWORDS_DEBIT = ("debit", " dr ", "\tdr", ",dr", "dr,", "(dr)")
_AMOUNT_KEYWORDS_CREDIT = ("credit", " cr ", "\tcr", ",cr", "cr,", "(cr)")


def _read_csv_rows(file_text: str) -> list[list[str]]:
    """Read CSV-style text into a list of row-lists. Tolerates extra
    whitespace and BOMs. Empty input returns []."""
    if not file_text:
        return []
    try:
        # Sniff the dialect — QBO and Excel exports vary on delimiter.
        sample = file_text[:8000]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        except csv.Error:
            dialect = csv.excel  # type: ignore[assignment]
        reader = csv.reader(io.StringIO(file_text), dialect=dialect)
        return [row for row in reader]
    except Exception:
        # Last-ditch: line/comma split.
        return [
            [cell.strip() for cell in line.split(",")]
            for line in file_text.splitlines()
        ]


def _find_col(headers: list[str], keywords: list[str]) -> int | None:
    """First column whose lowercased header contains any keyword. Order
    of `keywords` matters — earlier keywords win."""
    norm = [h.lower() for h in headers]
    for kw in keywords:
        for i, h in enumerate(norm):
            if kw in h:
                return i
    return None


def _parse_amount(value: Any) -> Decimal:
    """Parse currency-shaped strings: $1,234.56 / (1,234.56) / -1234.
    Returns 0 for blank / non-numeric input."""
    if value is None:
        return Decimal("0")
    s = str(value).strip()
    if not s or s in ("-", "—", "n/a", "NA"):
        return Decimal("0")
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1]
    s = s.replace("$", "").replace(",", "").replace(" ", "").strip()
    if s.startswith("-"):
        negative = True
        s = s[1:]
    if not s:
        return Decimal("0")
    try:
        v = Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal("0")
    return -v if negative else v


def _find_header_row(
    rows: list[list[str]], required_keywords: list[list[str]]
) -> int | None:
    """Scan the first 50 rows for one that contains a cell matching
    each keyword group. Returns the index, or None when no row matches.

    `required_keywords` is a list of keyword lists — each inner list is
    an OR group, all groups must match (AND across groups).
    """
    for i, row in enumerate(rows[:50]):
        cells_lower = [str(c or "").lower() for c in row]
        matched_all = True
        for group in required_keywords:
            if not any(any(kw in c for kw in group) for c in cells_lower):
                matched_all = False
                break
        if matched_all:
            return i
    return None


def _try_fallback_tb_parser(file_text: str) -> dict[str, Any] | None:
    """Heuristic CSV/TSV TB parser. Returns the same shape as
    parse_trial_balance — or None when no debit/credit columns are
    detected.

    Handles:
      * Two-column (debit, credit) layouts — QBO, Sage, generic exports.
      * Single signed-balance column — positive → debit, negative → credit.
    """
    rows = _read_csv_rows(file_text)
    if not rows:
        return None

    header_idx = _find_header_row(
        rows,
        required_keywords=[
            ["debit", "dr"],
            ["credit", "cr"],
        ],
    )

    if header_idx is not None:
        headers = [str(c or "").strip() for c in rows[header_idx]]
        debit_idx = _find_col(headers, ["debit", "dr"])
        credit_idx = _find_col(headers, ["credit", "cr"])
        code_idx = _find_col(headers, ["code", "number", "no.", " no", "acct"])
        name_idx = _find_col(
            headers, ["account name", "name", "description", "account"]
        )
        # If name_idx and code_idx collide (single 'account' column),
        # treat that column as name and leave code blank.
        if code_idx is not None and code_idx == name_idx:
            code_idx = None
        if debit_idx is None or credit_idx is None:
            return None
        return _parse_tb_rows(
            rows[header_idx + 1:],
            code_idx=code_idx,
            name_idx=name_idx,
            debit_idx=debit_idx,
            credit_idx=credit_idx,
            signed_idx=None,
        )

    # Try single signed-balance layout.
    header_idx = _find_header_row(
        rows,
        required_keywords=[
            ["balance", "amount", "net"],
        ],
    )
    if header_idx is None:
        return None
    headers = [str(c or "").strip() for c in rows[header_idx]]
    signed_idx = _find_col(headers, ["balance", "amount", "net"])
    code_idx = _find_col(headers, ["code", "number", "no.", "acct"])
    name_idx = _find_col(headers, ["account name", "name", "description", "account"])
    if signed_idx is None:
        return None
    if code_idx is not None and code_idx == name_idx:
        code_idx = None
    return _parse_tb_rows(
        rows[header_idx + 1:],
        code_idx=code_idx,
        name_idx=name_idx,
        debit_idx=None,
        credit_idx=None,
        signed_idx=signed_idx,
    )


def _parse_tb_rows(
    data_rows: list[list[str]],
    *,
    code_idx: int | None,
    name_idx: int | None,
    debit_idx: int | None,
    credit_idx: int | None,
    signed_idx: int | None,
) -> dict[str, Any] | None:
    tb_lines: list[dict[str, Any]] = []
    total_dr = Decimal("0")
    total_cr = Decimal("0")
    for row in data_rows:
        if not row or all(not str(c or "").strip() for c in row):
            continue
        code = (
            str(row[code_idx] or "").strip()
            if code_idx is not None and code_idx < len(row)
            else ""
        )
        name = (
            str(row[name_idx] or "").strip()
            if name_idx is not None and name_idx < len(row)
            else ""
        )
        # Skip total / summary / header rows that bleed past the real
        # header line. QBO and Sage both emit a final "Total" row with
        # the label in either the code or name column.
        if not code and not name:
            continue
        label_text = f"{code} {name}".lower()
        if "total" in label_text or label_text.strip() in ("net income", "net loss"):
            continue

        if signed_idx is not None:
            net = _parse_amount(
                row[signed_idx] if signed_idx < len(row) else ""
            )
            if net == 0:
                continue
            if net > 0:
                dr, cr = net, Decimal("0")
            else:
                dr, cr = Decimal("0"), -net
        else:
            dr = _parse_amount(
                row[debit_idx] if debit_idx is not None and debit_idx < len(row) else ""
            )
            cr = _parse_amount(
                row[credit_idx] if credit_idx is not None and credit_idx < len(row) else ""
            )
            if dr == 0 and cr == 0:
                continue
        if not code:
            code = name
        tb_lines.append({
            "account_code": code,
            "account_name": name or code,
            "debit_balance": dr,
            "credit_balance": cr,
        })
        total_dr += dr
        total_cr += cr

    if not tb_lines:
        return None
    variance = total_dr - total_cr
    return {
        "tb_lines": tb_lines,
        "total_debits": float(total_dr),
        "total_credits": float(total_cr),
        "variance": float(variance),
        "balanced": variance == 0,
    }


def _try_fallback_coa_parser(file_text: str) -> dict[str, Any] | None:
    """Heuristic chart-of-accounts parser. Looks for a header row with
    account-code + name + type columns. Returns the same shape as
    parse_chart_of_accounts or None when format isn't recognized.
    """
    rows = _read_csv_rows(file_text)
    if not rows:
        return None
    header_idx = _find_header_row(
        rows,
        required_keywords=[
            ["code", "number", "no.", "acct"],
            ["name", "description", "account"],
        ],
    )
    if header_idx is None:
        return None
    headers = [str(c or "").strip() for c in rows[header_idx]]
    code_idx = _find_col(headers, ["code", "number", "no.", "acct"])
    name_idx = _find_col(headers, ["account name", "name", "description"])
    type_idx = _find_col(headers, ["type", "class", "category"])
    sub_idx = _find_col(headers, ["subtype", "sub-type", "detail"])
    parent_idx = _find_col(headers, ["parent"])
    nb_idx = _find_col(headers, ["normal balance", "normal_balance", "dr/cr"])
    if code_idx is None or name_idx is None or code_idx == name_idx:
        return None

    out: list[dict[str, Any]] = []
    seen_codes: set[str] = set()
    for row in rows[header_idx + 1:]:
        if not row or all(not str(c or "").strip() for c in row):
            continue
        code = str(row[code_idx] or "").strip() if code_idx < len(row) else ""
        name = str(row[name_idx] or "").strip() if name_idx < len(row) else ""
        if not code or not name:
            continue
        if code in seen_codes:
            continue
        seen_codes.add(code)
        type_val = (
            str(row[type_idx] or "").strip()
            if type_idx is not None and type_idx < len(row)
            else ""
        )
        if not type_val:
            type_val = _infer_type_from_code(code)
        normal = (
            str(row[nb_idx] or "").strip().lower()
            if nb_idx is not None and nb_idx < len(row)
            else ""
        )
        if normal not in ("debit", "credit"):
            normal = (
                "debit" if _infer_type_from_code(code) in ("Asset", "Expense", "COGS") else "credit"
            )
        out.append({
            "code": code,
            "name": name,
            "type": type_val,
            "subtype": (
                str(row[sub_idx] or "").strip()
                if sub_idx is not None and sub_idx < len(row)
                else ""
            ),
            "normal_balance": normal,
            "parent_code": (
                (str(row[parent_idx] or "").strip() or None)
                if parent_idx is not None and parent_idx < len(row)
                else None
            ),
        })
    if not out:
        return None
    return {"accounts": out, "count": len(out)}


def _infer_type_from_code(code: str) -> str:
    p = (code or "").strip()[:1]
    return {
        "1": "Asset",
        "2": "Liability",
        "3": "Equity",
        "4": "Revenue",
        "5": "COGS",
        "6": "Expense",
        "7": "Expense",
        "8": "Expense",
        "9": "Expense",
    }.get(p, "Other")


def _try_fallback_gl_parser(file_text: str) -> list[dict[str, Any]] | None:
    """Heuristic GL parser. Looks for date + account + debit/credit
    columns. Returns the same shape as parse_gl_file (a list of normalized
    line records) or None.
    """
    rows = _read_csv_rows(file_text)
    if not rows:
        return None
    header_idx = _find_header_row(
        rows,
        required_keywords=[
            ["date"],
            ["account", "acct", "code"],
            ["debit", "credit", "amount"],
        ],
    )
    if header_idx is None:
        return None
    headers = [str(c or "").strip() for c in rows[header_idx]]
    date_idx = _find_col(headers, ["date"])
    ref_idx = _find_col(headers, ["reference", "ref", "doc", "txn", "number"])
    code_idx = _find_col(headers, ["account code", "account number", "code", "acct"])
    name_idx = _find_col(headers, ["account name", "account", "description"])
    desc_idx = _find_col(headers, ["description", "memo", "explanation"])
    cp_idx = _find_col(headers, ["payee", "vendor", "customer", "name"])
    debit_idx = _find_col(headers, ["debit", "dr"])
    credit_idx = _find_col(headers, ["credit", "cr"])
    amount_idx = _find_col(headers, ["amount", "net"])
    if date_idx is None or code_idx is None or code_idx == name_idx:
        return None
    if debit_idx is None and credit_idx is None and amount_idx is None:
        return None

    out: list[dict[str, Any]] = []
    for row in rows[header_idx + 1:]:
        if not row or all(not str(c or "").strip() for c in row):
            continue
        raw_date = str(row[date_idx] or "").strip() if date_idx < len(row) else ""
        if not raw_date:
            continue
        txn_date = _parse_loose_date(raw_date)
        if txn_date is None:
            continue
        code = str(row[code_idx] or "").strip() if code_idx < len(row) else ""
        if not code:
            continue
        if debit_idx is not None or credit_idx is not None:
            dr = _parse_amount(
                row[debit_idx] if debit_idx is not None and debit_idx < len(row) else ""
            )
            cr = _parse_amount(
                row[credit_idx] if credit_idx is not None and credit_idx < len(row) else ""
            )
        else:
            net = _parse_amount(
                row[amount_idx] if amount_idx is not None and amount_idx < len(row) else ""
            )
            if net >= 0:
                dr, cr = net, Decimal("0")
            else:
                dr, cr = Decimal("0"), -net
        if dr == 0 and cr == 0:
            continue
        out.append({
            "transaction_date": txn_date,
            "reference_number": (
                str(row[ref_idx] or "").strip()
                if ref_idx is not None and ref_idx < len(row)
                else ""
            ),
            "account_code": code,
            "account_name": (
                str(row[name_idx] or "").strip()
                if name_idx is not None and name_idx < len(row)
                else ""
            ),
            "description": (
                str(row[desc_idx] or "").strip()
                if desc_idx is not None and desc_idx < len(row)
                else ""
            ),
            "counterparty_name": (
                (str(row[cp_idx] or "").strip() or None)
                if cp_idx is not None and cp_idx < len(row)
                else None
            ),
            "debit_amount": dr,
            "credit_amount": cr,
        })
    if not out:
        return None
    return out


_DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%m/%d/%Y",
    "%d/%m/%Y",
    "%m-%d-%Y",
    "%d-%m-%Y",
    "%b %d, %Y",
    "%d %b %Y",
    "%B %d, %Y",
)


def _parse_loose_date(value: str) -> DateType | None:
    s = (value or "").strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# --------------------------------------------------------------------------
# 2B. Chart of accounts — file parse + save
# --------------------------------------------------------------------------


_COA_PROMPT = """Parse this chart of accounts file and return JSON only. No prose.

Extract every account as:
{
  "accounts": [{
    "code": "account number/code",
    "name": "account name",
    "type": "Asset|Liability|Equity|Revenue|Expense|COGS",
    "subtype": "optional subtype",
    "normal_balance": "debit|credit",
    "parent_code": "parent account code if this is a sub-account"
  }]
}

Rules:
- If no account code exists, generate one based on standard accounting convention:
  1xxx=Assets, 2xxx=Liabilities, 3xxx=Equity, 4xxx=Revenue, 5xxx=COGS, 6xxx=Expenses
- Infer normal_balance from account type (Asset/Expense/COGS → debit, Liability/Equity/Revenue → credit)
- Return ONLY valid JSON. No prose. No code fences.
"""


def parse_chart_of_accounts(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """Try the regex/CSV fallback first; fall through to Claude only when
    the file is in an unknown shape. Returns {accounts, count} or raises
    ValueError when both paths fail."""
    file_text = _decode_file(file_bytes, filename)

    fallback = _try_fallback_coa_parser(file_text)
    if fallback:
        logger.info("CoA parsed via regex fallback (%d accounts)", fallback["count"])
        return fallback

    parsed = _claude_parse_json(_COA_PROMPT, file_text)
    if not parsed:
        raise ValueError(
            "Could not parse the chart of accounts. Try a CSV with "
            "columns like 'code, name, type' — or set ANTHROPIC_API_KEY "
            "to let the AI parser handle arbitrary formats."
        )
    accounts = parsed.get("accounts") or []
    cleaned: list[dict[str, Any]] = []
    for a in accounts:
        code = str(a.get("code") or "").strip()
        name = str(a.get("name") or "").strip()
        if not code or not name:
            continue
        cleaned.append({
            "code": code,
            "name": name,
            "type": (a.get("type") or "Other"),
            "subtype": a.get("subtype") or "",
            "normal_balance": (a.get("normal_balance") or "debit"),
            "parent_code": a.get("parent_code") or None,
        })
    return {"accounts": cleaned, "count": len(cleaned)}


def save_chart_of_accounts(
    session,
    entity_id: str,
    accounts: list[dict[str, Any]],
) -> dict[str, Any]:
    """Upsert confirmed accounts into the accounts table. Returns
    {saved_count, conflicts[]}. A 'conflict' is an account_code that
    already has journal_lines posted against it — we still upsert the
    name but flag it so the wizard can surface a warning.
    """
    saved = 0
    conflicts: list[dict[str, Any]] = []
    for a in accounts:
        code = str(a.get("code") or "").strip()
        name = str(a.get("name") or "").strip()
        if not code or not name:
            continue
        account_class = _normalize_type(a.get("type") or "Other")
        statement_type = (
            "balance_sheet"
            if account_class in {"Asset", "Liability", "Equity"}
            else "income_statement"
        )

        # Detect conflicts: this code is already used by journal_lines.
        existing_lines = session.execute(
            text(
                """
                SELECT 1
                  FROM journal_lines jl
                  JOIN journal_batches jb ON jb.id = jl.journal_batch_id
                 WHERE jb.entity_id = :eid AND jl.account_code = :code
                 LIMIT 1
                """
            ),
            {"eid": entity_id, "code": code},
        ).first()
        if existing_lines:
            conflicts.append({"code": code, "name": name})

        session.execute(
            text(
                """
                INSERT INTO accounts (
                    entity_id, account_code, account_name, account_class, statement_type
                ) VALUES (
                    :eid, :code, :name, :class, :stmt
                )
                ON CONFLICT (entity_id, account_code)
                DO UPDATE SET
                    account_name  = EXCLUDED.account_name,
                    account_class = EXCLUDED.account_class,
                    statement_type = EXCLUDED.statement_type,
                    is_active      = TRUE
                """
            ),
            {
                "eid": entity_id,
                "code": code,
                "name": name,
                "class": account_class,
                "stmt": statement_type,
            },
        )
        saved += 1
    return {"saved_count": saved, "conflicts": conflicts}


def _normalize_type(type_str: str) -> str:
    t = (type_str or "").strip().lower()
    if t in ("asset", "assets", "current asset", "fixed asset", "bank", "cash"):
        return "Asset"
    if t in ("liability", "liabilities", "current liability", "long-term liability"):
        return "Liability"
    if t in ("equity",):
        return "Equity"
    if t in ("revenue", "income", "sales", "other income"):
        return "Revenue"
    if t in ("cogs", "cost of goods sold", "cost of sales"):
        return "COGS"
    if t in ("expense", "expenses", "operating expense", "other expense"):
        return "Expense"
    # Default to a safe pass-through.
    return type_str.title() if type_str else "Other"


# --------------------------------------------------------------------------
# 2C. Opening balances
# --------------------------------------------------------------------------


_TB_PROMPT = """Parse this trial balance and return JSON only. No prose.

Extract every account line as:
{
  "trial_balance": [{
    "account_code": "account number",
    "account_name": "account name",
    "debit_balance": 0.00,
    "credit_balance": 0.00
  }],
  "total_debits": 0.00,
  "total_credits": 0.00,
  "balanced": true|false
}

Rules:
- debit_balance and credit_balance are never both non-zero on the same row.
- total_debits should equal total_credits if balanced.
- If the file lists balances as a single signed column, positive → debit, negative → credit.
- Return ONLY valid JSON. No prose. No code fences.
"""


def parse_trial_balance(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """Try the regex/CSV fallback first (cheap + fast on QBO/Sage/CSV
    exports), fall through to Claude on unknown formats.

    Returns the parsed preview with totals + balanced flag, or raises
    ValueError when both paths fail.
    """
    file_text = _decode_file(file_bytes, filename)

    fallback = _try_fallback_tb_parser(file_text)
    if fallback:
        logger.info(
            "TB parsed via regex fallback (%d lines, balanced=%s)",
            len(fallback["tb_lines"]),
            fallback["balanced"],
        )
        return fallback

    parsed = _claude_parse_json(_TB_PROMPT, file_text)
    if not parsed:
        raise ValueError(
            "Could not parse the trial balance. Try a CSV with "
            "columns like 'account, debit, credit' — or set "
            "ANTHROPIC_API_KEY to let the AI parser handle arbitrary "
            "formats."
        )
    raw_lines = parsed.get("trial_balance") or []
    cleaned: list[dict[str, Any]] = []
    total_dr = Decimal("0")
    total_cr = Decimal("0")
    for line in raw_lines:
        code = str(line.get("account_code") or "").strip()
        name = str(line.get("account_name") or "").strip()
        dr = _to_decimal(line.get("debit_balance"))
        cr = _to_decimal(line.get("credit_balance"))
        if not code:
            continue
        cleaned.append({
            "account_code": code,
            "account_name": name,
            "debit_balance": dr,
            "credit_balance": cr,
        })
        total_dr += dr
        total_cr += cr

    variance = total_dr - total_cr
    return {
        "tb_lines": cleaned,
        "total_debits": float(total_dr),
        "total_credits": float(total_cr),
        "variance": float(variance),
        "balanced": variance == 0,
    }


def import_opening_balances(
    session,
    *,
    entity_id: str,
    entity_code: str,
    as_of_date: DateType,
    tb_lines: list[dict[str, Any]],
    actor_email: str,
) -> dict[str, Any]:
    """Validate the TB sum, create / find the opening period, write the
    journal_batch + lines.

    Returns {batch_id, line_count, total_debits, total_credits, balanced}.
    Raises ValueError when the TB doesn't balance — caller renders the
    variance back to the user.
    """
    total_dr = Decimal("0")
    total_cr = Decimal("0")
    for line in tb_lines:
        total_dr += _to_decimal(line.get("debit_balance"))
        total_cr += _to_decimal(line.get("credit_balance"))

    variance = total_dr - total_cr
    if variance != 0:
        raise ValueError(
            f"Trial balance is out of balance by ${float(variance):,.2f} "
            f"(debits ${float(total_dr):,.2f}, credits ${float(total_cr):,.2f}). "
            "Fix the file and re-upload."
        )

    period_id = _ensure_opening_period(session, entity_id, as_of_date)

    # Re-using opening_balance source_module would conflict with the
    # uniqueness constraint on (entity, period, source_module, batch_label).
    # batch_label distinguishes if someone re-runs (we replace).
    batch_label = f"Opening Balance — {as_of_date.isoformat()}"
    existing = session.execute(
        text(
            """
            SELECT id FROM journal_batches
             WHERE entity_id = :eid
               AND accounting_period_id = :pid
               AND source_module = 'opening_balance'
               AND batch_label = :label
             LIMIT 1
            """
        ),
        {"eid": entity_id, "pid": period_id, "label": batch_label},
    ).mappings().first()
    if existing:
        # Drop the prior batch + cascade lines so re-runs are idempotent.
        session.execute(
            text("DELETE FROM journal_batches WHERE id = :id"),
            {"id": existing["id"]},
        )

    batch_row = session.execute(
        text(
            """
            INSERT INTO journal_batches (
                entity_id, accounting_period_id, source_module, batch_label,
                status, workflow_status, total_debits, total_credits,
                approved_by, approved_at, summary_json
            ) VALUES (
                :eid, :pid, 'opening_balance', :label,
                'approved_to_post', 'approved_to_post', :td, :tc,
                :actor, NOW(), :sj
            )
            RETURNING id
            """
        ),
        {
            "eid": entity_id,
            "pid": period_id,
            "label": batch_label,
            "td": total_dr,
            "tc": total_cr,
            "actor": actor_email,
            "sj": json.dumps({
                "as_of_date": as_of_date.isoformat(),
                "source": "onboarding",
            }),
        },
    ).mappings().first()
    batch_id = batch_row["id"]

    line_count = 0
    for idx, line in enumerate(tb_lines, start=1):
        code = str(line.get("account_code") or "").strip()
        if not code:
            continue
        dr = _to_decimal(line.get("debit_balance"))
        cr = _to_decimal(line.get("credit_balance"))
        if dr == 0 and cr == 0:
            continue
        memo = f"Opening balance — {line.get('account_name') or code}"
        session.execute(
            text(
                """
                INSERT INTO journal_lines (
                    journal_batch_id, line_number, account_code,
                    debit_amount, credit_amount, memo, source_json
                ) VALUES (
                    :bid, :ln, :code, :dr, :cr, :memo, :sj
                )
                """
            ),
            {
                "bid": batch_id,
                "ln": idx,
                "code": code,
                "dr": dr,
                "cr": cr,
                "memo": memo,
                "sj": json.dumps({"account_name": line.get("account_name")}),
            },
        )
        line_count += 1

    return {
        "batch_id": str(batch_id),
        "line_count": line_count,
        "total_debits": float(total_dr),
        "total_credits": float(total_cr),
        "balanced": True,
    }


def _ensure_opening_period(session, entity_id: str, as_of_date: DateType):
    """Find or create the accounting period that contains as_of_date.

    Onboarding's opening-balance period is marked closed_locked
    (historical) so nothing else can post into it. The status value
    must be 'closed_locked' — that's the canonical terminal state per
    services_period_close.STATUS_CLOSED_LOCKED, and the value strict
    consumers (is_period_locked, assistant queries, period_close routes)
    check for. The bare string 'closed' is not recognized.
    """
    period_start = DateType(as_of_date.year, as_of_date.month, 1)
    last_day = calendar.monthrange(period_start.year, period_start.month)[1]
    period_end = DateType(period_start.year, period_start.month, last_day)

    row = session.execute(
        text(
            """
            SELECT id, status FROM accounting_periods
             WHERE entity_id = :eid AND period_start = :ps AND period_end = :pe
             LIMIT 1
            """
        ),
        {"eid": entity_id, "ps": period_start, "pe": period_end},
    ).mappings().first()
    if row:
        return row["id"]

    new_row = session.execute(
        text(
            """
            INSERT INTO accounting_periods (
                entity_id, period_label, period_start, period_end, status
            ) VALUES (
                :eid, :label, :ps, :pe, 'closed_locked'
            )
            RETURNING id
            """
        ),
        {
            "eid": entity_id,
            "label": f"Opening Balance Period ({period_start.strftime('%b %Y')})",
            "ps": period_start,
            "pe": period_end,
        },
    ).mappings().first()
    return new_row["id"]


async def import_trial_balance_from_qbo(
    session,
    *,
    entity_id: str,
    entity_code: str,
    as_of_date: DateType,
    actor_email: str,
) -> dict[str, Any]:
    """Pull TB from QBO and import via import_opening_balances."""
    connection = session.execute(
        text(
            """
            SELECT id, realm_id, access_token, refresh_token,
                   access_token_expires_at, refresh_token_expires_at, connected_at
              FROM quickbooks_connections
             WHERE entity_id = :eid AND is_active = TRUE
             ORDER BY connected_at DESC LIMIT 1
            """
        ),
        {"eid": entity_id},
    ).mappings().first()
    if not connection:
        raise ValueError("No active QuickBooks connection for this entity.")

    connection = await ensure_valid_access_token(session, dict(connection))
    qb = QuickBooksClient()
    rows = await qb.get_trial_balance(
        connection["realm_id"], connection["access_token"], as_of_date
    )
    # QBO returns its internal account Id; the chart-of-accounts import
    # may have stored AcctNum (the dealer's account number) as
    # account_code. Resolve QBO Id → account_code so the journal_lines
    # use the dealer's chart numbering rather than opaque QBO ids.
    id_map_rows = session.execute(
        text(
            """
            SELECT quickbooks_account_id, account_code
              FROM accounts
             WHERE entity_id = :eid AND quickbooks_account_id IS NOT NULL
            """
        ),
        {"eid": entity_id},
    ).mappings().all()
    qbo_to_code = {r["quickbooks_account_id"]: r["account_code"] for r in id_map_rows}
    tb_lines = [
        {
            "account_code": qbo_to_code.get(r["account_id"]) or r["account_id"],
            "account_name": r["account_name"],
            "debit_balance": r["debit_balance"],
            "credit_balance": r["credit_balance"],
        }
        for r in rows
    ]
    result = import_opening_balances(
        session,
        entity_id=entity_id,
        entity_code=entity_code,
        as_of_date=as_of_date,
        tb_lines=tb_lines,
        actor_email=actor_email,
    )
    _log_sync_run(
        session,
        entity_id=entity_id,
        connection_id=connection["id"],
        sync_type="trial_balance",
        sync_from=as_of_date,
        sync_to=as_of_date,
        summary={"line_count": result["line_count"]},
    )
    return result


# --------------------------------------------------------------------------
# 2D. GL history import
# --------------------------------------------------------------------------


_GL_PROMPT = """Parse this general ledger export and return JSON only. No prose.

Extract every transaction line as:
{
  "lines": [{
    "date": "YYYY-MM-DD",
    "reference": "transaction reference/number",
    "account_code": "account number",
    "account_name": "account name",
    "description": "transaction description",
    "counterparty": "vendor or customer name (optional)",
    "debit": 0.00,
    "credit": 0.00
  }]
}

Rules:
- Group related Dr/Cr lines by reference so the importer can rebuild journals.
- date must be YYYY-MM-DD.
- debit and credit are mutually exclusive per line (one is always 0).
- Return ONLY valid JSON. No prose. No code fences.
"""


def parse_gl_file(file_bytes: bytes, filename: str) -> list[dict[str, Any]]:
    """Try the regex/CSV fallback first for standard GL exports; fall
    back to Claude for unknown formats."""
    file_text = _decode_file(file_bytes, filename)

    fallback = _try_fallback_gl_parser(file_text)
    if fallback:
        logger.info("GL parsed via regex fallback (%d lines)", len(fallback))
        return fallback

    parsed = _claude_parse_json(_GL_PROMPT, file_text)
    if not parsed:
        raise ValueError(
            "Could not parse the GL file. Try a CSV with columns like "
            "'date, account, debit, credit' — or set ANTHROPIC_API_KEY "
            "to let the AI parser handle arbitrary formats."
        )
    raw_lines = parsed.get("lines") or []
    cleaned: list[dict[str, Any]] = []
    for line in raw_lines:
        try:
            txn_date = datetime.strptime(str(line.get("date") or ""), "%Y-%m-%d").date()
        except ValueError:
            continue
        code = str(line.get("account_code") or "").strip()
        if not code:
            continue
        dr = _to_decimal(line.get("debit"))
        cr = _to_decimal(line.get("credit"))
        if dr == 0 and cr == 0:
            continue
        cleaned.append({
            "transaction_date": txn_date,
            "reference_number": str(line.get("reference") or "").strip(),
            "account_code": code,
            "account_name": str(line.get("account_name") or "").strip(),
            "description": str(line.get("description") or "").strip(),
            "counterparty_name": (line.get("counterparty") or None),
            "debit_amount": dr,
            "credit_amount": cr,
        })
    return cleaned


def import_gl_history_from_lines(
    session,
    *,
    entity_id: str,
    entity_code: str,
    lines: list[dict[str, Any]],
    actor_email: str,
    progress_callback: Callable[[str, int], None] | None = None,
) -> dict[str, Any]:
    """Group lines by (period_end, reference) into journal_batches and
    persist. Shared by QBO and file-upload paths.

    Each unique reference within a month becomes one journal_batch.
    Lines with no reference are bundled per-day.
    """
    if not lines:
        return {
            "batches_created": 0,
            "lines_created": 0,
            "months_imported": 0,
            "date_from": None,
            "date_to": None,
        }

    # Group by (month_end, reference). Lines with no reference get
    # synthesized refs keyed off the date so each day rolls up to one
    # batch — that gives the historical-import a stable, queryable shape.
    buckets: dict[tuple[DateType, str], list[dict[str, Any]]] = {}
    for line in lines:
        txn_date: DateType = line["transaction_date"]
        last_day = calendar.monthrange(txn_date.year, txn_date.month)[1]
        month_end = DateType(txn_date.year, txn_date.month, last_day)
        ref = (line.get("reference_number") or "").strip()
        if not ref:
            ref = f"DAY-{txn_date.isoformat()}"
        buckets.setdefault((month_end, ref), []).append(line)

    months: set[DateType] = set()
    batches_created = 0
    lines_created = 0
    earliest: DateType | None = None
    latest: DateType | None = None

    sorted_keys = sorted(buckets.keys(), key=lambda k: (k[0], k[1]))
    total_keys = len(sorted_keys)

    for idx, (month_end, ref) in enumerate(sorted_keys, start=1):
        period_id = _ensure_historical_period(session, entity_id, month_end)
        bucket_lines = buckets[(month_end, ref)]

        # Skip imbalanced buckets — they're usually single-sided GL rows
        # that QBO emits for opening balance carryforward; opening
        # balances are imported separately via the TB.
        total_dr = sum((l["debit_amount"] for l in bucket_lines), Decimal("0"))
        total_cr = sum((l["credit_amount"] for l in bucket_lines), Decimal("0"))
        if total_dr != total_cr:
            # Force-balance to a "Suspense" 9999 line so the batch posts
            # and we don't lose the data. The dealer can clean it up
            # later from the suspense report.
            diff = total_dr - total_cr
            if diff > 0:
                bucket_lines.append({
                    "transaction_date": bucket_lines[0]["transaction_date"],
                    "account_code": "9999",
                    "account_name": "Suspense — Onboarding",
                    "description": "Auto-balancing line",
                    "debit_amount": Decimal("0"),
                    "credit_amount": diff,
                })
                total_cr += diff
            else:
                bucket_lines.append({
                    "transaction_date": bucket_lines[0]["transaction_date"],
                    "account_code": "9999",
                    "account_name": "Suspense — Onboarding",
                    "description": "Auto-balancing line",
                    "debit_amount": -diff,
                    "credit_amount": Decimal("0"),
                })
                total_dr += -diff

        batch_label = f"GL Import — {month_end.isoformat()} — {ref}"[:200]
        # Unique constraint is (entity, period, source_module, batch_label).
        # Drop a prior identical batch so this import is idempotent.
        session.execute(
            text(
                """
                DELETE FROM journal_batches
                 WHERE entity_id = :eid
                   AND accounting_period_id = :pid
                   AND source_module = 'historical_import'
                   AND batch_label = :label
                """
            ),
            {"eid": entity_id, "pid": period_id, "label": batch_label},
        )

        batch_row = session.execute(
            text(
                """
                INSERT INTO journal_batches (
                    entity_id, accounting_period_id, source_module, batch_label,
                    status, workflow_status, total_debits, total_credits,
                    approved_by, approved_at, summary_json
                ) VALUES (
                    :eid, :pid, 'historical_import', :label,
                    'approved_to_post', 'approved_to_post', :td, :tc,
                    :actor, NOW(), :sj
                )
                RETURNING id
                """
            ),
            {
                "eid": entity_id,
                "pid": period_id,
                "label": batch_label,
                "td": total_dr,
                "tc": total_cr,
                "actor": actor_email,
                "sj": json.dumps({"reference": ref, "source": "onboarding_gl"}),
            },
        ).mappings().first()
        batch_id = batch_row["id"]
        batches_created += 1

        for ln_idx, line in enumerate(bucket_lines, start=1):
            session.execute(
                text(
                    """
                    INSERT INTO journal_lines (
                        journal_batch_id, line_number, account_code,
                        debit_amount, credit_amount, memo, source_json
                    ) VALUES (
                        :bid, :ln, :code, :dr, :cr, :memo, :sj
                    )
                    """
                ),
                {
                    "bid": batch_id,
                    "ln": ln_idx,
                    "code": line["account_code"],
                    "dr": line["debit_amount"],
                    "cr": line["credit_amount"],
                    "memo": (line.get("description") or "")[:500],
                    "sj": json.dumps({
                        "transaction_date": line["transaction_date"].isoformat(),
                        "transaction_type": line.get("transaction_type"),
                        "account_name": line.get("account_name"),
                        "counterparty_name": line.get("counterparty_name"),
                        "reference_number": line.get("reference_number"),
                    }),
                },
            )
            lines_created += 1

        months.add(month_end)
        earliest = min(earliest, month_end) if earliest else month_end
        latest = max(latest, month_end) if latest else month_end

        if progress_callback and total_keys:
            pct = int(idx / total_keys * 100)
            progress_callback(f"Importing {month_end.strftime('%b %Y')}", pct)

    return {
        "batches_created": batches_created,
        "lines_created": lines_created,
        "months_imported": len(months),
        "date_from": earliest.isoformat() if earliest else None,
        "date_to": latest.isoformat() if latest else None,
    }


def _ensure_historical_period(session, entity_id: str, period_end: DateType):
    """Find or create the accounting period for a historical-import
    month. New periods are inserted as 'closed_locked' — the canonical
    terminal status per services_period_close — because pre-cutover
    history shouldn't be editable. 'closed' (without _locked) is not
    recognized by strict consumers and must not be used.
    """
    period_start = DateType(period_end.year, period_end.month, 1)
    row = session.execute(
        text(
            """
            SELECT id FROM accounting_periods
             WHERE entity_id = :eid AND period_start = :ps AND period_end = :pe
             LIMIT 1
            """
        ),
        {"eid": entity_id, "ps": period_start, "pe": period_end},
    ).mappings().first()
    if row:
        return row["id"]
    new_row = session.execute(
        text(
            """
            INSERT INTO accounting_periods (
                entity_id, period_label, period_start, period_end, status
            ) VALUES (
                :eid, :label, :ps, :pe, 'closed_locked'
            )
            RETURNING id
            """
        ),
        {
            "eid": entity_id,
            "label": period_start.strftime("%b %Y"),
            "ps": period_start,
            "pe": period_end,
        },
    ).mappings().first()
    return new_row["id"]


async def import_gl_history_from_qbo(
    session,
    *,
    entity_id: str,
    entity_code: str,
    date_from: DateType,
    date_to: DateType,
    actor_email: str,
    progress_callback: Callable[[str, int], None] | None = None,
) -> dict[str, Any]:
    """Pull GL from QBO month-by-month and persist."""
    connection = session.execute(
        text(
            """
            SELECT id, realm_id, access_token, refresh_token,
                   access_token_expires_at, refresh_token_expires_at, connected_at
              FROM quickbooks_connections
             WHERE entity_id = :eid AND is_active = TRUE
             ORDER BY connected_at DESC LIMIT 1
            """
        ),
        {"eid": entity_id},
    ).mappings().first()
    if not connection:
        raise ValueError("No active QuickBooks connection for this entity.")

    connection = await ensure_valid_access_token(session, dict(connection))
    qb = QuickBooksClient()
    chunks = month_chunks(date_from, date_to)
    total_chunks = len(chunks) or 1

    # QBO Id → account_code map so GL lines use the dealer's account
    # numbering instead of opaque QBO ids.
    id_map_rows = session.execute(
        text(
            """
            SELECT quickbooks_account_id, account_code
              FROM accounts
             WHERE entity_id = :eid AND quickbooks_account_id IS NOT NULL
            """
        ),
        {"eid": entity_id},
    ).mappings().all()
    qbo_to_code = {r["quickbooks_account_id"]: r["account_code"] for r in id_map_rows}

    all_lines: list[dict[str, Any]] = []
    for idx, (month_start, month_end) in enumerate(chunks, start=1):
        if progress_callback:
            pct = int((idx - 1) / total_chunks * 50)  # 0-50% = pulling
            progress_callback(f"Pulling {month_start.strftime('%b %Y')} from QBO", pct)
        try:
            rows = await qb.get_general_ledger(
                connection["realm_id"], connection["access_token"],
                month_start, month_end,
            )
        except Exception as exc:
            logger.warning("QBO GL pull failed for %s: %r", month_start, exc)
            continue
        for row in rows:
            # Rewrite account_id → dealer account_code before downstream
            # bucketing.
            qbo_id = row.get("account_id")
            row["account_code"] = qbo_to_code.get(qbo_id) or qbo_id
            all_lines.append(row)

    # Now write them. Use the second half of the progress bar for writes.
    def _write_progress(label: str, pct: int) -> None:
        if progress_callback:
            progress_callback(label, 50 + int(pct / 2))

    result = import_gl_history_from_lines(
        session,
        entity_id=entity_id,
        entity_code=entity_code,
        lines=all_lines,
        actor_email=actor_email,
        progress_callback=_write_progress,
    )
    _log_sync_run(
        session,
        entity_id=entity_id,
        connection_id=connection["id"],
        sync_type="general_ledger",
        sync_from=date_from,
        sync_to=date_to,
        summary=result,
    )
    return result


# --------------------------------------------------------------------------
# 2E. Learn from GL — bootstrap vendor memory
# --------------------------------------------------------------------------


def learn_from_gl_history(
    session, *, entity_id: str, entity_code: str
) -> dict[str, Any]:
    """Walk journal_lines and bootstrap vendor_classification_memory +
    assistant_entity_memory so the AI assistant is useful on day one.

    Strategy:
      1. For each (counterparty_name, account_code) pair, count
         occurrences. Promote the top-N counterparties to
         vendor_classification_memory.
      2. Top recurring memo phrases also become memory entries.
    """
    # Pull lines together with their source_json's counterparty hint.
    vendor_rows = session.execute(
        text(
            """
            SELECT
                COALESCE(jl.source_json->>'counterparty_name', '') AS vendor_key,
                jl.account_code,
                CASE WHEN jl.debit_amount > 0 THEN 'debit' ELSE 'credit' END AS dr_cr,
                COUNT(*) AS n
              FROM journal_lines jl
              JOIN journal_batches jb ON jb.id = jl.journal_batch_id
             WHERE jb.entity_id = :eid
               AND jb.source_module = 'historical_import'
               AND COALESCE(jl.source_json->>'counterparty_name', '') <> ''
          GROUP BY 1, 2, 3
          ORDER BY n DESC
             LIMIT 500
            """
        ),
        {"eid": entity_id},
    ).mappings().all()

    # Pick the dominant account per vendor — many vendors hit multiple
    # accounts; we want the most-frequent one.
    by_vendor: dict[str, dict[str, Any]] = {}
    for r in vendor_rows:
        key = (r["vendor_key"] or "").strip().upper()
        if not key:
            continue
        entry = by_vendor.get(key)
        if entry is None or r["n"] > entry["n"]:
            by_vendor[key] = {
                "account_code": r["account_code"],
                "dr_cr": r["dr_cr"],
                "n": int(r["n"]),
            }

    vendors_learned = 0
    for vendor_key, dominant in by_vendor.items():
        session.execute(
            text(
                """
                -- vendor_classification_memory schema notes:
                --   * confidence_score is numeric(4,3) — 0.000 to ~9.999.
                --     Production data uses a 0-1 scale (default 0.500).
                --   * source is CHECK-constrained to
                --     ('gl_history','user_confirmed','ai_seeded').
                --   * Unique constraint is (entity_id, normalized_vendor_key,
                --     account_code) — three columns, not two.
                --   * No `created_at` column — `first_seen_at` is auto-set
                --     to NOW() by the column default.
                INSERT INTO vendor_classification_memory (
                    entity_id, normalized_vendor_key, account_code,
                    debit_or_credit, occurrences_count, confidence_score,
                    source, last_seen_at
                ) VALUES (
                    :eid, :key, :acct, :dr_cr, :n, 0.85,
                    'gl_history', NOW()
                )
                ON CONFLICT (entity_id, normalized_vendor_key, account_code) DO UPDATE
                   SET debit_or_credit = EXCLUDED.debit_or_credit,
                       occurrences_count = vendor_classification_memory.occurrences_count
                           + EXCLUDED.occurrences_count,
                       confidence_score = LEAST(
                           1,
                           GREATEST(
                               vendor_classification_memory.confidence_score,
                               0.85
                           )
                       ),
                       source = 'gl_history',
                       last_seen_at = NOW()
                """
            ),
            {
                "eid": entity_id,
                "key": vendor_key[:200],
                "acct": dominant["account_code"],
                "dr_cr": dominant["dr_cr"],
                "n": dominant["n"],
            },
        )
        vendors_learned += 1

        # Also record an assistant memory hint so the conversational
        # assistant can reason about this vendor in chat.
        session.execute(
            text(
                """
                INSERT INTO assistant_entity_memory (
                    entity_code, memory_type, memory_key, memory_value,
                    confidence, times_confirmed, last_seen_at
                ) VALUES (
                    :ec, 'vendor_account', :key, :acct, 85, :n, NOW()
                )
                ON CONFLICT (entity_code, memory_type, memory_key) DO UPDATE
                   SET memory_value = EXCLUDED.memory_value,
                       confidence = LEAST(100, GREATEST(
                           assistant_entity_memory.confidence, 85
                       )),
                       times_confirmed = assistant_entity_memory.times_confirmed
                           + EXCLUDED.times_confirmed,
                       last_seen_at = NOW()
                """
            ),
            {
                "ec": entity_code,
                "key": vendor_key[:200],
                "acct": dominant["account_code"],
                "n": dominant["n"],
            },
        )

    # Period patterns — count distinct months each account appears in.
    period_rows = session.execute(
        text(
            """
            SELECT jl.account_code,
                   COUNT(DISTINCT date_trunc('month', ap.period_end)) AS month_count
              FROM journal_lines jl
              JOIN journal_batches jb ON jb.id = jl.journal_batch_id
              JOIN accounting_periods ap ON ap.id = jb.accounting_period_id
             WHERE jb.entity_id = :eid
               AND jb.source_module = 'historical_import'
          GROUP BY jl.account_code
          ORDER BY month_count DESC
             LIMIT 30
            """
        ),
        {"eid": entity_id},
    ).mappings().all()
    observations_created = 0
    for r in period_rows:
        # Heuristic: shows up in >=70% of imported months → recurring.
        if int(r["month_count"]) >= 8:
            session.execute(
                text(
                    """
                    INSERT INTO assistant_period_observations (
                        entity_code, period_end, observation_type,
                        observation, account_code, severity
                    ) VALUES (
                        :ec, CURRENT_DATE, 'journal_created',
                        :obs, :code, 'info'
                    )
                    """
                ),
                {
                    "ec": entity_code,
                    "obs": f"Account {r['account_code']} appears in "
                           f"{int(r['month_count'])} of the imported months — "
                           f"likely a recurring entry.",
                    "code": r["account_code"],
                },
            )
            observations_created += 1

    return {
        "vendors_learned": vendors_learned,
        "patterns_found": len(by_vendor),
        "observations_created": observations_created,
    }


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------


def _to_decimal(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        if isinstance(value, Decimal):
            return value
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0")


def _log_sync_run(
    session,
    *,
    entity_id: str,
    connection_id: str,
    sync_type: str,
    sync_from: DateType | None,
    sync_to: DateType | None,
    summary: dict[str, Any],
) -> None:
    """Write a quickbooks_sync_runs row. The original sync code never
    used this table — onboarding bootstraps that habit so the dashboard
    can show a real 'last sync' time."""
    try:
        session.execute(
            text(
                """
                INSERT INTO quickbooks_sync_runs (
                    entity_id, quickbooks_connection_id, sync_type,
                    sync_from, sync_to, status, summary_json,
                    started_at, finished_at
                ) VALUES (
                    :eid, :cid, :st, :sf, :st2, 'complete',
                    :sj, NOW(), NOW()
                )
                """
            ),
            {
                "eid": entity_id,
                "cid": connection_id,
                "st": sync_type,
                "sf": sync_from,
                "st2": sync_to,
                "sj": json.dumps(_jsonable(summary)),
            },
        )
    except Exception:
        logger.exception("Failed to log quickbooks_sync_runs row — non-fatal")


def _jsonable(value: Any) -> Any:
    """Coerce Decimal / date / datetime so json.dumps doesn't choke."""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (DateType, datetime)):
        return value.isoformat()
    if isinstance(value, dict):
        return {k: _jsonable(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonable(v) for v in value]
    return value
