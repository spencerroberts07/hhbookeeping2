"""
Analytics-only daily-sales backfill (Phase 2A).

Parses the Bridlewood "Cash Balancing" workbooks (weekly tabs, days as columns)
and inserts one row per business_date into `cash_balancing_days` with the daily
`total_sales` ("Item Sales"), `total_hst` ("Tax - HST"), and `opening_cash`.

IMPORTANT: this is READ-ONLY with respect to accounting data. It writes ONLY to
`cash_balancing_days` (the daily sales source table) — it NEVER creates
`journal_batches` / `journal_lines`, so closed-period GL is untouched. Existing
rows (the real Feb–May 2026 imports) are protected via ON CONFLICT DO NOTHING.

Usage (from backend/, venv python):
    python scripts/backfill_daily_sales.py --entity 1877-8 \
        "C:/Users/.../Cash Balancing FY25 - Bridlewood HH.xlsx" \
        "C:/Users/.../Cash Balancing FY26 - Bridlewood HH (2).xlsx"
    python scripts/backfill_daily_sales.py --entity 1877-8 ... --commit
Without --commit it runs a dry run and reports what WOULD be inserted.
"""
from __future__ import annotations

import argparse
import datetime
import json
import sys
from pathlib import Path

import openpyxl
from sqlalchemy import text

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from app.db import db_session  # noqa: E402


def _norm(v) -> str:
    return str(v).strip().lower() if v is not None else ""


def _find_row(ws, *labels: str) -> int | None:
    """Row index whose column-A label starts with any of `labels`."""
    for r in range(1, (ws.max_row or 0) + 1):
        cell = _norm(ws.cell(r, 1).value)
        if cell and any(cell.startswith(lbl) for lbl in labels):
            return r
    return None


def parse_workbook(path: str) -> dict[datetime.date, dict]:
    """Return {business_date: {item_sales, hst, opening_cash, tab}} for a file."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out: dict[datetime.date, dict] = {}
    fname = Path(path).name
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if not ws.max_row or ws.max_row < 4:
            continue
        # The day-header row holds the seven day-dates across cols 2..8. Pick
        # the row (within the first 8) with the MOST datetime cells, so we
        # don't mistake the "Period <start> to <end>" row (2 datetimes) for it.
        best_row, best_count = None, 0
        for r in range(1, 9):
            cnt = sum(1 for c in range(2, (ws.max_column or 0) + 1)
                      if isinstance(ws.cell(r, c).value, datetime.datetime))
            if cnt > best_count:
                best_row, best_count = r, cnt
        if best_row is None or best_count < 3:
            continue
        date_row = best_row
        sales_row = _find_row(ws, "item sales")
        hst_row = _find_row(ws, "tax - hst", "tax-hst")
        open_row = _find_row(ws, "opening cash")
        if sales_row is None:
            continue
        for c in range(2, (ws.max_column or 0) + 1):
            dv = ws.cell(date_row, c).value
            if not isinstance(dv, datetime.datetime):
                continue
            d = dv.date()
            sales = ws.cell(sales_row, c).value
            if not isinstance(sales, (int, float)) or sales <= 0:
                continue  # skip empty/template/closed days
            hst = ws.cell(hst_row, c).value if hst_row else None
            opening = ws.cell(open_row, c).value if open_row else None
            # First file to populate a date wins (handles FY25/FY26 overlap).
            out.setdefault(d, {
                "item_sales": round(float(sales), 2),
                "hst": round(float(hst), 2) if isinstance(hst, (int, float)) else None,
                "opening_cash": round(float(opening), 2) if isinstance(opening, (int, float)) else None,
                "tab": sheet,
                "source": fname,
            })
    wb.close()
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--entity", required=True)
    ap.add_argument("files", nargs="+")
    ap.add_argument("--commit", action="store_true")
    args = ap.parse_args()

    merged: dict[datetime.date, dict] = {}
    for f in args.files:
        parsed = parse_workbook(f)
        for d, v in parsed.items():
            merged.setdefault(d, v)
        print(f"parsed {Path(f).name}: {len(parsed)} dated days")

    today = datetime.date.today()
    merged = {d: v for d, v in merged.items() if d <= today}
    print(f"total candidate days (<= today): {len(merged)} "
          f"[{min(merged)} -> {max(merged)}]")

    with db_session() as s:
        eid = s.execute(text("SELECT id FROM entities WHERE entity_code=:e"),
                        {"e": args.entity}).scalar()
        if not eid:
            sys.exit(f"entity {args.entity} not found")
        periods = s.execute(text(
            "SELECT id, period_start, period_end FROM accounting_periods WHERE entity_id=:e"
        ), {"e": eid}).mappings().all()

        def period_for(d: datetime.date):
            for p in periods:
                if p["period_start"] <= d <= p["period_end"]:
                    return p["id"]
            return None

        existing = {r[0] for r in s.execute(text(
            "SELECT business_date FROM cash_balancing_days WHERE entity_id=:e"
        ), {"e": eid})}
        to_insert = {d: v for d, v in merged.items() if d not in existing}
        print(f"already present: {len(existing)} | to insert: {len(to_insert)}")

        if not args.commit:
            sample = sorted(to_insert.items())[:5]
            for d, v in sample:
                print(f"  would insert {d} sales={v['item_sales']} hst={v['hst']} tab={v['tab']}")
            print("DRY RUN — re-run with --commit to write.")
            return

        n = 0
        for d, v in sorted(to_insert.items()):
            s.execute(text(
                """
                INSERT INTO cash_balancing_days
                    (entity_id, accounting_period_id, business_date, tab_name,
                     total_sales, total_hst, opening_cash, raw_json)
                VALUES
                    (:eid, :pid, :bd, :tab, :sales, :hst, :opening, CAST(:rj AS jsonb))
                ON CONFLICT (entity_id, business_date) DO NOTHING
                """
            ), {
                "eid": eid, "pid": period_for(d), "bd": d, "tab": v["tab"],
                "sales": v["item_sales"], "hst": v["hst"], "opening": v["opening_cash"],
                "rj": json.dumps({"backfill": True, "source": v["source"],
                                  "item_sales": v["item_sales"]}),
            })
            n += 1
        print(f"inserted {n} daily rows (analytics-only, no GL posting)")


if __name__ == "__main__":
    main()
